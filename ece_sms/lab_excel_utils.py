import io
import os
import re
import statistics

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image

from models import Student, Enrollment
from calculations import validate_lab_component

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LEFT_LOGO = os.path.join(BASE_DIR, "static", "images", "left_logo.jpg")
RIGHT_LOGO = os.path.join(BASE_DIR, "static", "images", "right_logo.png")

TNR = "Times New Roman"
BLACK = "000000"
WHITE = "FFFFFF"
DARK_BLUE = "1E3A5F"
LIGHT_GRAY = "F2F3F4"
PASS_FILL = "D5F5E3"
FAIL_FILL = "FADBD8"
PENDING_FILL = "FFF3CD"


def _thin_border():
    s = Side(style="thin", color=BLACK)
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_border():
    s = Side(style="medium", color=BLACK)
    return Border(left=s, right=s, top=s, bottom=s)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _body_font(size=10, bold=False, color=BLACK):
    return Font(name=TNR, size=size, bold=bold, color=color)


def _sanitize(name):
    name = str(name or "").strip()
    name = re.sub(r"[^\w\s-]", "", name)
    name = re.sub(r"[\s]+", "_", name)
    return name or "Lab"


def _roll(sr):
    return f"EC31{sr:02d}"


def _add_common_header(ws, last_col_letter, right_logo_cell):
    for i in range(1, 6):
        ws.row_dimensions[i].height = 24
    ws.row_dimensions[1].height = 70

    if os.path.exists(LEFT_LOGO):
        left_logo = Image(LEFT_LOGO)
        left_logo.width = 95
        left_logo.height = 95
        ws.add_image(left_logo, "A1")

    if os.path.exists(RIGHT_LOGO):
        right_logo = Image(RIGHT_LOGO)
        right_logo.width = 95
        right_logo.height = 95
        ws.add_image(right_logo, right_logo_cell)

    ws.merge_cells(f"B1:{last_col_letter}1")
    ws["B1"] = "Chhatrapati Shahu Maharaj Shikshan Sanstha"
    ws.merge_cells(f"B2:{last_col_letter}2")
    ws["B2"] = "CSMSS Chh. Shahu College of Engineering"
    ws.merge_cells(f"B3:{last_col_letter}3")
    ws["B3"] = "Kanchanwadi, Chhatrapati Sambhajinagar - 431011 (Maharashtra)"
    ws.merge_cells(f"B4:{last_col_letter}5")
    ws["B4"] = "Department of Electronics and Computer Engineering"

    ws["B1"].font = _body_font(12)
    ws["B2"].font = _body_font(20, bold=True)
    ws["B3"].font = _body_font(11)
    ws["B4"].font = _body_font(18, bold=True)

    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).border = _thick_border()
            ws.cell(r, c).alignment = _center()


def generate_lab_template(meta, students):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lab Marks"

    widths = {
        "A": 8, "B": 12, "C": 16, "D": 34, "E": 13,
        "F": 13, "G": 14, "H": 14, "I": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    _add_common_header(ws, "H", "I1")

    row = 6
    ws[f"A{row}"] = "Subject Code :"
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"] = meta["subject_code"]
    ws[f"D{row}"] = "Subject Name :"
    ws.merge_cells(f"E{row}:I{row}")
    ws[f"E{row}"] = meta["subject_name"]

    row = 7
    ws[f"A{row}"] = "Semester :"
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"] = meta["semester_id"]
    ws[f"D{row}"] = "Academic Year :"
    ws.merge_cells(f"E{row}:I{row}")
    ws[f"E{row}"] = meta["academic_year"]

    row = 8
    ws[f"A{row}"] = "Formula :"
    ws.merge_cells(f"B{row}:I{row}")
    ws[f"B{row}"] = "CA1 /30 + CA2 /30 = Internal /60; External /40; Total /100" + (f" | {meta.get('batch_label', 'All Students')}" if meta.get("batch_label") else "")

    for r in range(6, 9):
        for c in range(1, 10):
            cell = ws.cell(r, c)
            cell.border = _thick_border()
            cell.font = _body_font(11, bold=(c in [1, 4]))
            cell.alignment = _left()

    header_row = 10
    headers = [
        "SR", "Roll No", "PRN", "Student Name",
        "CA1 /30", "CA2 /30", "Internal /60", "External /40", "Total /100"
    ]

    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(header_row, col, hdr)
        cell.font = _body_font(12, bold=True)
        cell.alignment = _center()
        cell.border = _thick_border()
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for sr, student in enumerate(students, 1):
        r = header_row + sr
        ca1 = student.get("ca1")
        ca2 = student.get("ca2")
        external = student.get("external")
        internal = student.get("internal")
        total = student.get("total_marks")

        values = [
            sr,
            _roll(sr),
            student["prn"],
            student["name"],
            ca1,
            ca2,
            internal,
            external,
            total,
        ]
        for c, value in enumerate(values, 1):
            cell = ws.cell(r, c, value)
            cell.font = _body_font(11)
            cell.border = _thin_border()
            cell.alignment = _center() if c != 4 else _left()
            if c in [1, 2, 3, 4, 7, 9]:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    note_row = header_row + len(students) + 1
    ws.merge_cells(f"A{note_row}:I{note_row}")
    ws[f"A{note_row}"] = "NOTE: Enter marks only in CA1, CA2 and External columns. Do not edit student details."
    ws[f"A{note_row}"].font = _body_font(11, bold=True)
    ws[f"A{note_row}"].alignment = _left()
    ws[f"A{note_row}"].border = _thick_border()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Lab_Template_{_sanitize(meta['subject_code'])}.xlsx"
    return buf, filename


def _parse_mark(value, max_marks, label, row_num, errors):
    valid, err = validate_lab_component(value, max_marks, label)
    if not valid:
        errors.append({"row": row_num, "col": label, "msg": err})
        return None
    return float(value)


def parse_lab_upload(file_obj, subject_code, semester_id, academic_year):
    try:
        wb = load_workbook(file_obj, data_only=True)
        ws = wb.active
    except Exception as e:
        return False, [{"row": 0, "col": "FILE", "msg": f"Cannot read file: {e}"}], 0

    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "Roll No" in str(cell.value):
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        return False, [{"row": 0, "col": "STRUCTURE", "msg": "Header row not found. Use the official template."}], 0

    valid_prns = {s.prn for s in Student.query.all()}
    enrolled_prns = {
        e.prn for e in Enrollment.query.filter_by(
            subject_code=subject_code,
            semester_id=semester_id,
            academic_year=academic_year
        ).all()
    }

    errors = []
    records = []

    for idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not row or len(row) < 8:
            continue
        if not row[2]:
            continue

        prn = str(row[2]).strip()
        if prn.lower().startswith("note"):
            continue

        if prn not in valid_prns:
            errors.append({"row": idx, "col": "PRN", "msg": f"PRN {prn} not found in database"})
            continue
        if prn not in enrolled_prns:
            errors.append({"row": idx, "col": "Enrollment", "msg": f"PRN {prn} is not enrolled in this lab/project subject"})
            continue

        ca1_val = row[4]
        ca2_val = row[5]
        ext_val = row[7]

        # Completely blank marks row = skipped/absent/pending.
        if ca1_val in [None, ""] and ca2_val in [None, ""] and ext_val in [None, ""]:
            records.append({"prn": prn, "absent": True})
            continue

        ca1 = _parse_mark(ca1_val, 30, "CA1", idx, errors)
        ca2 = _parse_mark(ca2_val, 30, "CA2", idx, errors)
        external = _parse_mark(ext_val, 40, "External", idx, errors)

        if ca1 is None or ca2 is None or external is None:
            continue

        records.append({
            "prn": prn,
            "ca1": ca1,
            "ca2": ca2,
            "external": external,
            "absent": False,
        })

    if errors:
        return False, errors, 0

    return True, records, len(records)


def generate_lab_excel_report(meta, data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lab Report"

    for col, width in {
        "A": 7, "B": 12, "C": 16, "D": 32, "E": 12, "F": 12,
        "G": 14, "H": 14, "I": 13, "J": 10, "K": 12, "L": 12
    }.items():
        ws.column_dimensions[col].width = width

    _add_common_header(ws, "K", "L1")

    ws.merge_cells("A6:L6")
    ws["A6"] = f"LAB / PROJECT MARKS REPORT | {meta['subject_code']} - {meta['subject_name']} | Semester {meta['semester_id']} | A.Y. {meta['academic_year']} | {meta.get('batch_label', 'All Students')}"
    ws["A6"].font = _body_font(12, bold=True)
    ws["A6"].alignment = _center()
    ws["A6"].border = _thick_border()

    headers = [
        "SR", "Roll No", "PRN", "Student Name", "CA1 /30", "CA2 /30",
        "Internal /60", "External /40", "Total /100", "Grade", "Grade Point", "Result"
    ]
    hr = 8
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(hr, col, hdr)
        cell.font = _body_font(10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = _center()
        cell.border = _thick_border()

    pass_c = fail_c = pending_c = 0
    totals = []

    for i, row in enumerate(data, 1):
        r = hr + i
        status = row["status"]
        if status == "PASS":
            fill = PASS_FILL
            pass_c += 1
        elif status == "FAIL":
            fill = FAIL_FILL
            fail_c += 1
        else:
            fill = PENDING_FILL
            pending_c += 1

        values = [
            i, _roll(i), row["prn"], row["name"], row.get("ca1"), row.get("ca2"),
            row.get("internal"), row.get("external"), row.get("total"), row.get("grade"),
            row.get("grade_point"), status
        ]

        if row.get("total") is not None:
            totals.append(float(row["total"]))

        for col, value in enumerate(values, 1):
            cell = ws.cell(r, col, value if value is not None else "-")
            cell.font = _body_font(10)
            cell.alignment = _center() if col != 4 else _left()
            cell.border = _thin_border()
            if col == 12:
                cell.fill = PatternFill("solid", fgColor=fill)

    sr = hr + len(data) + 2
    summary = [
        "SUMMARY",
        f"Total: {len(data)}",
        f"Pass: {pass_c}",
        f"Fail: {fail_c}",
        f"Pending: {pending_c}",
        f"Highest: {max(totals) if totals else '-'}",
        f"Average: {round(sum(totals)/len(totals), 2) if totals else '-'}",
    ]
    for col, val in enumerate(summary, 1):
        cell = ws.cell(sr, col, val)
        cell.font = _body_font(10, bold=True)
        cell.border = _thick_border()
        cell.alignment = _center()

    sig_row = sr + 3
    ws.row_dimensions[sig_row].height = 60
    ws.merge_cells(f"A{sig_row}:D{sig_row}")
    ws.merge_cells(f"E{sig_row}:H{sig_row}")
    ws.merge_cells(f"I{sig_row}:L{sig_row}")
    ws[f"A{sig_row}"] = "Lab Incharge\n(Signature)"
    ws[f"E{sig_row}"] = "Class Teacher\n(Signature)"
    ws[f"I{sig_row}"] = "HOD\n(Signature)"
    for cell_ref in [f"A{sig_row}", f"E{sig_row}", f"I{sig_row}"]:
        cell = ws[cell_ref]
        cell.font = _body_font(11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
        cell.border = _thick_border()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Lab_Report_{_sanitize(meta['subject_code'])}_Sem{meta['semester_id']}.xlsx"
    return buf, filename
