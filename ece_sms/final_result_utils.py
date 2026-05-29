from io import BytesIO
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from sqlalchemy import or_

from models import db, Student, Subject, Enrollment, TheoryMarks, LabMarks, ExternalMarks
from batch_utils import clean_filter, batch_label
from calculations import compute_final_result, compute_lab_result, get_grade, update_internal_totals, is_theory_internal_complete, is_gradable_subject


TNR = "Times New Roman"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LEFT_LOGO = os.path.join(BASE_DIR, "static", "images", "left_logo.jpg")
RIGHT_LOGO = os.path.join(BASE_DIR, "static", "images", "right_logo.png")


# -------------------------------------------------
# Basic Excel styles
# -------------------------------------------------
def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _thin_border():
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _header_fill():
    return PatternFill("solid", fgColor="D9EAF7")


def _title_fill():
    return PatternFill("solid", fgColor="1F4E78")


def _title_font():
    return Font(name=TNR, size=14, bold=True, color="FFFFFF")


def _header_font():
    return Font(name=TNR, size=11, bold=True, color="000000")


def _body_font(bold=False):
    return Font(name=TNR, size=10, bold=bold, color="000000")


def _safe_float(value, default=0.0):
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


# -------------------------------------------------
# Backward-compatible grade helper
# -------------------------------------------------
def calculate_grade(total_marks, external_marks=None):
    """
    Backward-compatible wrapper.

    THEORY result decisions are delegated to calculations.compute_final_result(),
    so the ESE minimum remains one single source of truth: external >= 20.
    """
    total = _safe_float(total_marks)

    if external_marks is not None:
        external = _safe_float(external_marks)
        internal = total - external
        result = compute_final_result(
            internal_marks=internal,
            external_marks=external,
            credits=0
        )
        return {
            "grade": result["grade"] or "FF",
            "grade_point": float(result["grade_point"] or 0),
            "is_passed": bool(result["is_passed"])
        }

    # Used only for LAB/PROJECT totals where no separate ESE minimum is applied.
    if total < 40:
        return {
            "grade": "FF",
            "grade_point": 0.0,
            "is_passed": False
        }

    grade, gp = get_grade(total)
    return {
        "grade": grade,
        "grade_point": float(gp),
        "is_passed": True
    }


# -------------------------------------------------
# Recalculate final result for one TheoryMarks row
# -------------------------------------------------
def recalculate_theory_result(mark_row):
    """
    Updates one TheoryMarks row using calculations.py as the single source of truth.

    Theory pass conditions:
    - Internal: /40
    - External: /60
    - Total: /100
    - Pass only when external >= 20 AND total >= 40

    Also refreshes internal_total using the confirmed formula:
    Internal = best 2 of (CT1, CT2, Assignment) + MSE.
    """

    if mark_row is None:
        return None

    # Keep internal fields consistent with the confirmed best-2-of-3 CA formula.
    update_internal_totals(mark_row)

    # Final result should remain PENDING until every internal component and
    # the external marks are present. This prevents incomplete records from
    # appearing as FAIL just because missing marks were treated as 0.
    if not is_theory_internal_complete(mark_row) or mark_row.external is None:
        mark_row.total_marks = None
        mark_row.grade = None
        mark_row.grade_point = None
        mark_row.is_passed = None
        return mark_row

    subject = Subject.query.filter_by(subject_code=mark_row.subject_code).first()
    credits = int(subject.credits or 0) if subject else 0

    result = compute_final_result(
        internal_marks=mark_row.internal_total,
        external_marks=mark_row.external,
        credits=credits
    )

    mark_row.total_marks = result["total"]
    mark_row.grade = result["grade"]
    mark_row.grade_point = result["grade_point"]
    mark_row.is_passed = result["is_passed"]

    return mark_row


# -------------------------------------------------
# Recalculate result for one LabMarks / Project row
# -------------------------------------------------
def recalculate_lab_result(lab_row):
    """
    Lab/Project result:
    - CA1: /30
    - CA2: /30
    - Internal: CA1 + CA2 = /60
    - External: /40
    - Total: /100
    - Pass when total >= 40

    Compatibility: if old rows have internal/external but no ca1/ca2,
    the existing internal value is still used.
    """

    if lab_row is None:
        return None

    if lab_row.external is None:
        lab_row.total_marks = None
        lab_row.grade = None
        lab_row.grade_point = None
        lab_row.is_passed = None
        return lab_row

    ca1 = getattr(lab_row, "ca1", None)
    ca2 = getattr(lab_row, "ca2", None)

    if ca1 is not None and ca2 is not None:
        subject = Subject.query.filter_by(subject_code=lab_row.subject_code).first()
        credits = int(subject.credits or 0) if subject else 0
        result = compute_lab_result(
            ca1_marks=ca1,
            ca2_marks=ca2,
            external_marks=lab_row.external,
            credits=credits
        )
        lab_row.internal = result["internal"]
        lab_row.total_marks = result["total"]
        lab_row.grade = result["grade"]
        lab_row.grade_point = result["grade_point"]
        lab_row.is_passed = result["is_passed"]
        return lab_row

    if lab_row.internal is None:
        lab_row.total_marks = None
        lab_row.grade = None
        lab_row.grade_point = None
        lab_row.is_passed = None
        return lab_row

    total = _safe_float(lab_row.internal) + _safe_float(lab_row.external)
    lab_row.total_marks = round(total, 2)

    if total < 40:
        lab_row.grade = "FF"
        lab_row.grade_point = 0.0
        lab_row.is_passed = False
    else:
        grade, gp = get_grade(total)
        lab_row.grade = grade
        lab_row.grade_point = gp
        lab_row.is_passed = True

    return lab_row


# -------------------------------------------------
# Internal helpers for report calculation
# -------------------------------------------------
def _add_subject_result(rows, subject, internal, external, total, grade, grade_point, result):
    credits = int(subject.credits or 0)
    credit_points = None

    if grade_point is not None:
        credit_points = round(credits * _safe_float(grade_point), 2)

    rows.append({
        "subject_code": subject.subject_code,
        "subject_name": subject.subject_name,
        "subject_type": subject.subject_type,
        "credits": credits,
        "internal": internal,
        "external": external,
        "total": total,
        "grade": grade,
        "grade_point": grade_point,
        "credit_points": credit_points,
        "result": result
    })

    return credit_points


# -------------------------------------------------
# Build one student's semester result
# -------------------------------------------------
def build_student_semester_result(prn, semester_id, academic_year):
    student = Student.query.filter_by(prn=prn).first()

    enrolled_subjects = (
        db.session.query(Enrollment, Subject)
        .join(Subject, Enrollment.subject_code == Subject.subject_code)
        .filter(
            Enrollment.prn == prn,
            Enrollment.semester_id == semester_id,
            Enrollment.academic_year == academic_year,
            Subject.is_audit == False,
            Subject.is_active == True,
            Subject.is_attendance_only == False,
            # Elective parent rows are placeholders only. Count selected options only.
            or_(Subject.is_elective == False, Subject.parent_subject_code.isnot(None))
        )
        .order_by(Subject.subject_code)
        .all()
    )

    rows = []
    total_credits = 0
    total_credit_points = 0
    credits_earned = 0
    is_complete = True
    has_failed = False

    for enrollment, subject in enrolled_subjects:
        if not is_gradable_subject(subject):
            continue

        credits = int(subject.credits or 0)
        total_credits += credits

        internal = None
        external = None
        total = None
        grade = "-"
        grade_point = None
        result = "PENDING"

        if subject.subject_type == "THEORY":
            marks = TheoryMarks.query.filter_by(
                prn=prn,
                subject_code=subject.subject_code,
                semester_id=semester_id,
                academic_year=academic_year
            ).first()

            ext_row = ExternalMarks.query.filter_by(
                prn=prn,
                subject_code=subject.subject_code,
                semester_id=semester_id,
                academic_year=academic_year
            ).first()

            external_val = None
            if marks and marks.external is not None:
                external_val = marks.external
            elif ext_row and ext_row.external_marks is not None:
                external_val = ext_row.external_marks

            # Show external marks even when internal marks are still pending, but
            # do not create a blank TheoryMarks row just for external marks.
            external = external_val

            if marks:
                update_internal_totals(marks)
                internal = marks.internal_total if is_theory_internal_complete(marks) else None

                if marks.external is None and external_val is not None:
                    marks.external = external_val

                if is_theory_internal_complete(marks) and external_val is not None:
                    recalculate_theory_result(marks)

                    if marks.is_passed is None or marks.total_marks is None:
                        is_complete = False
                    else:
                        internal = marks.internal_total
                        external = marks.external
                        total = marks.total_marks
                        grade = marks.grade or "-"
                        grade_point = marks.grade_point
                        result = "PASS" if marks.is_passed is True else "FAIL"

                        if marks.is_passed is True:
                            credits_earned += credits
                        else:
                            has_failed = True
                else:
                    is_complete = False
            else:
                is_complete = False

        else:
            lab = LabMarks.query.filter_by(
                prn=prn,
                subject_code=subject.subject_code,
                semester_id=semester_id,
                academic_year=academic_year
            ).first()

            if lab and lab.external is not None and (lab.internal is not None or (getattr(lab, "ca1", None) is not None and getattr(lab, "ca2", None) is not None)):
                recalculate_lab_result(lab)

                internal = lab.internal
                external = lab.external
                total = lab.total_marks
                grade = lab.grade or "-"
                grade_point = lab.grade_point
                if lab.is_passed is None or lab.total_marks is None:
                    is_complete = False
                else:
                    result = "PASS" if lab.is_passed is True else "FAIL"

                    if lab.is_passed is True:
                        credits_earned += credits
                    else:
                        has_failed = True
            else:
                is_complete = False

        credit_points = _add_subject_result(
            rows=rows,
            subject=subject,
            internal=internal,
            external=external,
            total=total,
            grade=grade,
            grade_point=grade_point,
            result=result
        )

        if credit_points is not None:
            total_credit_points += credit_points

    if not enrolled_subjects:
        is_complete = False

    sgpa = None
    if is_complete and total_credits > 0:
        sgpa = round(total_credit_points / total_credits, 2)

    return {
        "student": student,
        "prn": prn,
        "semester_id": semester_id,
        "academic_year": academic_year,
        "subjects": rows,
        "total_credits": total_credits,
        "credits_earned": credits_earned,
        "total_credit_points": round(total_credit_points, 2),
        "sgpa": sgpa,
        "is_complete": is_complete,
        "has_failed": has_failed,
        "result": "PASS" if is_complete and not has_failed else ("FAIL" if is_complete else "PENDING")
    }


# -------------------------------------------------
# Build semester result for all students
# -------------------------------------------------
def build_semester_result(semester_id, academic_year, division="", batch=""):
    division = clean_filter(division)
    batch = clean_filter(batch)

    students_query = (
        db.session.query(Student)
        .join(Enrollment, Enrollment.prn == Student.prn)
        .join(Subject, Enrollment.subject_code == Subject.subject_code)
        .filter(
            Enrollment.semester_id == semester_id,
            Enrollment.academic_year == academic_year,
            Subject.is_audit == False,
            Subject.is_active == True,
            Subject.is_attendance_only == False,
            # Do not pull students only because of an invalid elective parent enrollment.
            or_(Subject.is_elective == False, Subject.parent_subject_code.isnot(None))
        )
        .distinct()
    )

    if division:
        students_query = students_query.filter(Student.division == division)
    if batch:
        students_query = students_query.filter(Student.batch == batch)

    students = students_query.order_by(Student.prn).all()

    result_rows = []

    for student in students:
        sem_result = build_student_semester_result(
            prn=student.prn,
            semester_id=semester_id,
            academic_year=academic_year
        )

        result_rows.append({
            "prn": student.prn,
            "name": student.name,
            "total_credits": sem_result["total_credits"],
            "credits_earned": sem_result["credits_earned"],
            "total_credit_points": sem_result["total_credit_points"],
            "sgpa": sem_result["sgpa"],
            "result": sem_result["result"]
        })

    return result_rows


# -------------------------------------------------
# Calculate CGPA for one student up to selected semester
# -------------------------------------------------
def calculate_cgpa_for_student(prn, upto_semester_id=None):
    """
    Calculate CGPA only when every semester up to selected semester is complete.
    If any earlier semester is PENDING, CGPA is also PENDING/blank.
    """
    sem_query = (
        db.session.query(Enrollment.semester_id, Enrollment.academic_year)
        .join(Subject, Enrollment.subject_code == Subject.subject_code)
        .filter(
            Enrollment.prn == prn,
            Subject.is_audit == False,
            Subject.is_active == True,
            Subject.is_attendance_only == False,
            # Elective parent rows are placeholders and must not create CGPA dependencies.
            or_(Subject.is_elective == False, Subject.parent_subject_code.isnot(None))
        )
        .distinct()
        .order_by(Enrollment.semester_id)
    )

    if upto_semester_id is not None:
        sem_query = sem_query.filter(Enrollment.semester_id <= upto_semester_id)

    semesters = sem_query.all()
    if not semesters:
        return None

    total_credits = 0
    total_credit_points = 0

    for sem_id, ay in semesters:
        sem_result = build_student_semester_result(prn, sem_id, ay)
        if not sem_result["is_complete"] or sem_result["sgpa"] is None:
            return None

        credits = int(sem_result["total_credits"] or 0)
        if credits <= 0:
            return None

        total_credits += credits
        total_credit_points += credits * _safe_float(sem_result["sgpa"])

    if total_credits == 0:
        return None

    return round(total_credit_points / total_credits, 2)


# -------------------------------------------------
# Build SGPA + CGPA report
# -------------------------------------------------
def build_sgpa_cgpa_report(semester_id, academic_year, division="", batch=""):
    semester_rows = build_semester_result(semester_id, academic_year, division=division, batch=batch)

    final_rows = []

    for row in semester_rows:
        cgpa = calculate_cgpa_for_student(
            prn=row["prn"],
            upto_semester_id=semester_id
        )

        final_rows.append({
            "prn": row["prn"],
            "name": row["name"],
            "semester_id": semester_id,
            "academic_year": academic_year,
            "sgpa": row["sgpa"],
            "cgpa": cgpa,
            "total_credits": row["total_credits"],
            "credits_earned": row.get("credits_earned"),
            "total_credit_points": row["total_credit_points"],
            "result": row["result"]
        })

    return final_rows


# -------------------------------------------------
# Generate SGPA/CGPA Excel report
# -------------------------------------------------
def generate_sgpa_cgpa_excel_report(semester_id, academic_year, division="", batch=""):
    division = clean_filter(division)
    batch = clean_filter(batch)
    rows = build_sgpa_cgpa_report(semester_id, academic_year, division=division, batch=batch)
    report_filter_label = batch_label(division, batch)

    wb = Workbook()
    ws = wb.active
    ws.title = "SGPA CGPA Report"

    widths = [8, 18, 30, 12, 14, 14, 16, 16, 14]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Official common header with logos, matching other ECE SMS reports.
    for r in range(1, 6):
        ws.row_dimensions[r].height = 24
    ws.row_dimensions[1].height = 70

    if os.path.exists(LEFT_LOGO):
        left_logo = Image(LEFT_LOGO)
        left_logo.width = 85
        left_logo.height = 85
        ws.add_image(left_logo, "A1")

    if os.path.exists(RIGHT_LOGO):
        right_logo = Image(RIGHT_LOGO)
        right_logo.width = 85
        right_logo.height = 85
        ws.add_image(right_logo, "I1")

    ws.merge_cells("B1:H1")
    ws["B1"] = "Chhatrapati Shahu Maharaj Shikshan Sanstha's"
    ws.merge_cells("B2:H2")
    ws["B2"] = "CSMSS Chh. Shahu College of Engineering"
    ws.merge_cells("B3:H3")
    ws["B3"] = "Kanchanwadi, Chhatrapati Sambhajinagar – 431011"
    ws.merge_cells("B4:H5")
    ws["B4"] = "Department of Electronics and Computer Engineering"

    for row in range(1, 6):
        for col in range(1, 10):
            c = ws.cell(row=row, column=col)
            c.border = _thin_border()
            c.alignment = _center()

    ws["B1"].font = Font(name=TNR, size=12)
    ws["B2"].font = Font(name=TNR, size=18, bold=True)
    ws["B3"].font = Font(name=TNR, size=11)
    ws["B4"].font = Font(name=TNR, size=16, bold=True)

    ws.merge_cells("A6:I6")
    title = ws["A6"]
    title.value = "SGPA / CGPA REPORT"
    title.font = _title_font()
    title.alignment = _center()
    title.fill = _title_fill()
    title.border = _thin_border()

    ws.merge_cells("A7:I7")
    sub = ws["A7"]
    sub.value = f"Semester: {semester_id}     Academic Year: {academic_year}     Filter: {report_filter_label}"
    sub.font = _body_font(bold=True)
    sub.alignment = _center()
    sub.border = _thin_border()

    headers = [
        "SR",
        "PRN",
        "Student Name",
        "Semester",
        "SGPA",
        "CGPA",
        "Total Credits",
        "Credit Points",
        "Result",
    ]

    start_row = 9

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col)
        cell.value = header
        cell.font = _header_font()
        cell.alignment = _center()
        cell.fill = _header_fill()
        cell.border = _thin_border()

    for idx, row in enumerate(rows, start=1):
        r = start_row + idx
        values = [
            idx,
            row["prn"],
            row["name"],
            row["semester_id"],
            row["sgpa"] if row["sgpa"] is not None else "PENDING",
            row["cgpa"] if row["cgpa"] is not None else "PENDING",
            row["total_credits"],
            row["total_credit_points"],
            row["result"],
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col)
            cell.value = value
            cell.font = _body_font()
            cell.alignment = _center() if col != 3 else _left()
            cell.border = _thin_border()

    sig_row = start_row + len(rows) + 4
    ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=3)
    ws.merge_cells(start_row=sig_row, start_column=4, end_row=sig_row, end_column=6)
    ws.merge_cells(start_row=sig_row, start_column=7, end_row=sig_row, end_column=9)

    for cell_ref, label in [(f"A{sig_row}", "Result Coordinator"), (f"D{sig_row}", "HOD"), (f"G{sig_row}", "Principal")]:
        c = ws[cell_ref]
        c.value = label
        c.font = _body_font(bold=True)
        c.alignment = _center()

    for row in ws.iter_rows(min_row=sig_row, max_row=sig_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = Border(top=Side(style="medium", color="000000"))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filter_suffix = "" if report_filter_label == "All Students" else "_" + report_filter_label.replace(" ", "_").replace("|", "")
    filename = f"SGPA_CGPA_Report_Sem_{semester_id}_{academic_year}{filter_suffix}.xlsx"
    filename = filename.replace("/", "-")

    return output, filename
