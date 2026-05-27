import io
import re
from datetime import date, datetime, time, timedelta
from decimal import Decimal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from models import (
    db, Student, Subject, Enrollment, Attendance,
    TimetableSlot, AttendanceSession, AttendanceEntry,
)
from batch_utils import apply_student_batch_filters

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
DAY_LOOKUP = {d.lower(): d for d in DAY_NAMES}
DAY_LOOKUP.update({d[:3].lower(): d for d in DAY_NAMES})
VALID_SESSION_TYPES = {"THEORY", "LAB", "PROJECT", "AUDIT", "ATTENDANCE_ONLY", "TRAINING", "TUTORIAL"}
PRESENT_VALUES = {"P", "1", "Y", "YES", "PRESENT"}
ABSENT_VALUES = {"A", "0", "N", "NO", "ABSENT"}
EXCUSED_VALUES = {"L", "LEAVE", "O", "OD", "OFFICIAL", "MEDICAL", "M"}

TNR = "Times New Roman"
BLACK = "000000"
WHITE = "FFFFFF"
GOLD = "C9A84C"
DARK = "1A2B4A"
LGRAY = "F4F2EE"
MGRAY = "D9D9D9"
GREEN = "DCFCE7"
RED = "FEE2E2"
YELLOW = "FEF3C7"


def _border():
    s = Side(style="thin", color="999999")
    return Border(left=s, right=s, top=s, bottom=s)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _clean(value, default=""):
    if value is None:
        return default
    return str(value).strip()


def clean_code(value):
    return _clean(value).upper().replace(" ", "")


def clean_day(value):
    raw = _clean(value).lower()
    return DAY_LOOKUP.get(raw)


def clean_filter_value(value, default="ALL"):
    raw = _clean(value, default).upper()
    if raw in {"", "-", "NONE", "NA", "N/A", "ALL"}:
        return "ALL"
    return raw


def clean_session_type(value):
    raw = _clean(value, "THEORY").upper().replace(" ", "_").replace("-", "_")
    return raw if raw in VALID_SESSION_TYPES else "THEORY"


def parse_time_value(value):
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    raw = raw.replace(".", ":")
    # Excel may return values like 10:00 AM, 10.00AM, 14:00
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M%p", "%I %p"):
        try:
            return datetime.strptime(raw.upper(), fmt).time().replace(second=0, microsecond=0)
        except ValueError:
            pass
    m = re.match(r"^(\d{1,2})\s*[:.]\s*(\d{2})$", raw)
    if m:
        h, minute = int(m.group(1)), int(m.group(2))
        if 0 <= h <= 23 and 0 <= minute <= 59:
            return time(h, minute)
    return None


def parse_date(value):
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    raw = _clean(value)
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    return None


def _apply_fill(cell, color):
    cell.fill = PatternFill('solid', fgColor=color)


def _style_cell(cell, *, bold=False, size=10, color=BLACK, fill=None, align='center'):
    cell.font = Font(name=TNR, size=size, bold=bold, color=color)
    cell.alignment = _center() if align == 'center' else _left()
    cell.border = _border()
    if fill:
        cell.fill = PatternFill('solid', fgColor=fill)


def generate_timetable_template():
    """
    Teacher-friendly timetable upload template.

    This is intentionally designed like the college class timetable layout:
    days on the left, time-slots on top, subject entries inside cells.

    Accepted cell formats:
    - BTECPC601 - Internet of Things
    - B1: BTECPL606 - IoT + AIML Lab [LAB]
    - APTITUDE6 - Aptitude Training [ATTENDANCE_ONLY]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Class Timetable"

    max_col = 9  # A:I

    # Header similar to the provided timetable format.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(1, 1).value = "CSMSS Chh. Shahu College of Engineering"
    _style_cell(ws.cell(1, 1), bold=True, size=14, color=WHITE, fill=DARK)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(2, 1).value = "DEPARTMENT OF ELECTRONICS & COMPUTER ENGINEERING"
    _style_cell(ws.cell(2, 1), bold=True, size=12, color=BLACK, fill=LGRAY)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
    ws.cell(3, 1).value = "CLASS TIME TABLE — WEEKLY ATTENDANCE SETUP"
    _style_cell(ws.cell(3, 1), bold=True, size=13, color=WHITE, fill=GOLD)

    # Editable metadata row.
    meta = [
        ("A4", "ACADEMIC YEAR"), ("B4", "2025-26"),
        ("C4", "SEMESTER"), ("D4", 6),
        ("E4", "CLASS"), ("F4", "T.Y. ECE"),
        ("G4", "DIVISION"), ("H4", "A"),
        ("I4", "Edit only subject cells below"),
    ]
    for addr, value in meta:
        ws[addr] = value
        _style_cell(ws[addr], bold=addr[0] in {'A','C','E','G'}, fill=LGRAY if addr[0] in {'A','C','E','G'} else None)

    # Time-table grid.
    headers = [
        "DAY / TIME",
        "10:00-11:15",
        "11:15-12:15",
        "12:15-01:00\nLUNCH BREAK",
        "01:00-02:00",
        "02:00-03:00",
        "03:00-03:15\nTEA BREAK",
        "03:15-04:15",
        "04:15-05:15",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=h)
        _style_cell(cell, bold=True, size=10, color=WHITE, fill=DARK)

    days = ["MON", "TUE", "WED", "THU", "FRI", "SAT"]
    samples = {
        "MON": {
            2: "BTECPC601 - Internet of Things",
            3: "BTECPC602 - AIML",
            5: "BTECPE603A - Professional Elective [THEORY]",
            6: "BTECOE604A - Open Elective [THEORY]",
            8: "BTECHM605A - Development Engineering",
        },
        "TUE": {
            2: "B1: BTECPL606 - IoT + AIML Lab [LAB]",
            3: "B2: BTECPL606 - IoT + AIML Lab [LAB]",
            5: "BTECPC601 - Internet of Things",
            8: "APTITUDE6 - Aptitude Training [ATTENDANCE_ONLY]",
            9: "SOFTSKILL6 - Soft Skill Training [ATTENDANCE_ONLY]",
        },
        "WED": {
            2: "BTECPC602 - AIML",
            3: "BTECPE603A - Professional Elective [THEORY]",
            5: "B1: BTECM607 - Mini Project II [PROJECT]",
            6: "B2: BTECM607 - Mini Project II [PROJECT]",
            8: "BTECOE604A - Open Elective [THEORY]",
        },
        "THU": {
            2: "BTECPC601 - Internet of Things",
            3: "BTECHM605A - Development Engineering",
            5: "BTECPC602 - AIML",
            8: "BTECPE603A - Professional Elective [THEORY]",
        },
        "FRI": {
            2: "BTECOE604A - Open Elective [THEORY]",
            3: "BTECPC601 - Internet of Things",
            5: "B1: BTECPL606 - IoT + AIML Lab [LAB]",
            6: "B2: BTECPL606 - IoT + AIML Lab [LAB]",
            8: "LIBRARY / MENTORING [ATTENDANCE_ONLY]",
        },
        "SAT": {
            2: "BTECHM605A - Development Engineering",
            3: "BTECPE603A - Professional Elective [THEORY]",
            5: "BTECOE604A - Open Elective [THEORY]",
            8: "SPORTS6 - Sports / Activity [ATTENDANCE_ONLY]",
        },
    }

    for row, day in enumerate(days, 7):
        ws.cell(row=row, column=1, value=day)
        _style_cell(ws.cell(row, 1), bold=True, size=11, color=WHITE, fill=GOLD)
        for col in range(2, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if col in (4, 7):
                cell.value = "LUNCH BREAK" if col == 4 else "TEA BREAK"
                _style_cell(cell, bold=True, size=9, color=BLACK, fill=MGRAY)
            else:
                cell.value = samples.get(day, {}).get(col, "")
                _style_cell(cell, size=9, align='center')

    # Notes/faculty helper area.
    start_note = 15
    ws.merge_cells(start_row=start_note, start_column=1, end_row=start_note, end_column=max_col)
    ws.cell(start_note, 1).value = "HOW TO EDIT SUBJECT CELLS"
    _style_cell(ws.cell(start_note, 1), bold=True, size=12, color=WHITE, fill=DARK)

    notes = [
        "Write exact Subject Code first. Example: BTECPC601 - Internet of Things",
        "For batch-wise labs, use prefix like B1: BTECPL606 - IoT + AIML Lab [LAB]",
        "For sessions counted only in attendance, use [ATTENDANCE_ONLY]. Example: APTITUDE6 - Aptitude Training [ATTENDANCE_ONLY]",
        "Leave Lunch/Tea cells unchanged. Empty cells are ignored.",
        "Semester, Academic Year and Division are taken from row 4.",
    ]
    for i, note in enumerate(notes, start_note + 1):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=max_col)
        ws.cell(i, 1).value = note
        _style_cell(ws.cell(i, 1), size=10, align='left')

    # Legacy flat upload sheet kept for power users / debugging.
    flat = wb.create_sheet("Timetable Upload")
    flat.sheet_state = "hidden"
    flat_headers = [
        "Semester", "Academic Year", "Division", "Batch", "Day", "Start Time",
        "End Time", "Subject Code", "Subject Name", "Session Type", "Faculty", "Room"
    ]
    for col, h in enumerate(flat_headers, 1):
        cell = flat.cell(row=4, column=col, value=h)
        _style_cell(cell, bold=True, color=WHITE, fill=GOLD)

    widths = [14, 18, 22, 18, 22, 22, 18, 22, 22]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for r in range(1, 22):
        ws.row_dimensions[r].height = 26
    for r in range(7, 13):
        ws.row_dimensions[r].height = 62
    ws.freeze_panes = "B7"

    note_ws = wb.create_sheet("Instructions")
    note_ws["A1"] = "Instructions"
    note_ws["A1"].font = Font(name=TNR, size=14, bold=True)
    instructions = [
        "1. Use the Class Timetable sheet. It looks like a normal college timetable.",
        "2. Edit subject cells only. Keep the time headers and day names unchanged.",
        "3. Use exact subject codes from Subject Management for theory/lab/project subjects.",
        "4. Add [ATTENDANCE_ONLY] for activities like Aptitude, Soft Skill, Library, Sports, Mentoring, etc.",
        "5. Use B1:, B2:, B3: prefix for batch-wise labs/practicals.",
        "6. The old row-wise Timetable Upload sheet is hidden and kept only for compatibility.",
    ]
    for i, line in enumerate(instructions, 3):
        note_ws.cell(row=i, column=1, value=line).font = Font(name=TNR, size=11)
    note_ws.column_dimensions["A"].width = 120

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, "Class_Timetable_Upload_Template.xlsx"


_TIME_RANGE_RE = re.compile(
    r"(\d{1,2}\s*[:.]\s*\d{2}\s*(?:AM|PM|am|pm)?)\s*(?:-|to|TO|–|—)\s*(\d{1,2}\s*[:.]\s*\d{2}\s*(?:AM|PM|am|pm)?)"
)
_SUBJECT_CODE_RE = re.compile(r"\b([A-Z]{2,}[A-Z0-9]*\d+[A-Z]?)\b")
_BATCH_PREFIX_RE = re.compile(r"^\s*(?:BATCH\s*)?([A-Z]?\d|B\d|T\d|ALL)\s*[:\-]\s*(.+)$", re.IGNORECASE)


def _parse_time_range(value):
    raw = _clean(value)
    if not raw:
        return None, None
    raw = raw.replace("\n", " ")
    m = _TIME_RANGE_RE.search(raw)
    if not m:
        return None, None
    return parse_time_value(m.group(1)), parse_time_value(m.group(2))


def _is_break_text(value):
    raw = _clean(value).upper()
    if not raw:
        return True
    return any(word in raw for word in ["BREAK", "LUNCH", "TEA", "RECESS"])


def _split_visual_cell_entries(value):
    raw = _clean(value)
    if not raw or _is_break_text(raw):
        return []
    lines = [ln.strip() for ln in raw.replace(";", "\n").splitlines() if ln.strip()]
    # If several lines have subject codes, treat each line as a separate timetable entry.
    coded_lines = [ln for ln in lines if _SUBJECT_CODE_RE.search(ln)]
    if len(coded_lines) >= 2:
        return coded_lines
    return [" ".join(lines)]


def _session_type_from_text(text, subject=None):
    raw = _clean(text).upper()
    if "ATTENDANCE_ONLY" in raw or "ATTENDANCE ONLY" in raw:
        return "ATTENDANCE_ONLY"
    if "TRAINING" in raw or "APTITUDE" in raw or "SOFT" in raw or "MENTOR" in raw or "LIBRARY" in raw or "SPORT" in raw:
        return "ATTENDANCE_ONLY"
    if "TUTORIAL" in raw:
        return "TUTORIAL"
    if "PROJECT" in raw or "MINI PROJECT" in raw:
        return "PROJECT"
    if "LAB" in raw or "PRACTICAL" in raw:
        return "LAB"
    if subject is not None and getattr(subject, 'subject_type', None):
        return subject.subject_type
    return "THEORY"


def _parse_visual_entry(text):
    raw = _clean(text)
    if not raw or _is_break_text(raw):
        return None

    batch = "ALL"
    m_batch = _BATCH_PREFIX_RE.match(raw)
    if m_batch:
        batch = clean_filter_value(m_batch.group(1))
        raw = m_batch.group(2).strip()

    m_code = _SUBJECT_CODE_RE.search(raw.upper())
    if m_code:
        subject_code = clean_code(m_code.group(1))
        name_part = raw[m_code.end():]
        name_part = re.sub(r"^[\s\-:—–]+", "", name_part).strip()
    else:
        # For attendance-only entries without code, generate a stable code from text.
        words = re.sub(r"[^A-Za-z0-9]+", "", raw.upper())[:16] or "ATTENDANCE"
        subject_code = words
        name_part = raw

    subject = Subject.query.get(subject_code)
    subject_name = subject.subject_name if subject and getattr(subject, 'subject_name', None) else name_part
    if not subject_name:
        subject_name = subject_code

    session_type = _session_type_from_text(raw, subject)
    faculty = ""
    for part in re.split(r"[|,]", raw):
        if "PROF" in part.upper() or "DR." in part.upper() or "MR." in part.upper() or "MS." in part.upper():
            faculty = part.strip()
            break

    return {
        "subject_code": subject_code,
        "subject_name": subject_name,
        "batch": batch,
        "session_type": clean_session_type(session_type),
        "faculty": faculty,
    }


def _ensure_subject_for_timetable(subject_code, subject_name, semester_id, session_type):
    subject = Subject.query.get(subject_code)
    attendance_only = session_type in {'ATTENDANCE_ONLY', 'TRAINING', 'TUTORIAL'}
    if not subject:
        subject = Subject(
            subject_code=subject_code,
            subject_name=subject_name,
            semester_id=semester_id,
            subject_type='AUDIT' if attendance_only else ('LAB' if session_type == 'LAB' else 'PROJECT' if session_type == 'PROJECT' else 'THEORY'),
            credits=0,
            l_hours=0,
            t_hours=0,
            p_hours=0,
            category='ATTEND' if attendance_only else '',
            is_elective=False,
            is_audit=True if attendance_only else False,
            is_attendance_only=True if attendance_only else False,
            is_active=True,
        )
        db.session.add(subject)
        db.session.flush()
    else:
        if attendance_only:
            subject.is_attendance_only = True
            subject.is_audit = True
            subject.credits = 0
            subject.category = subject.category or 'ATTEND'
        if not subject.subject_name:
            subject.subject_name = subject_name
    return subject


def _upsert_timetable_slot(semester_id, academic_year, division, batch, day, start, end, subject_code, subject_name, session_type, faculty=None, room=None):
    _ensure_subject_for_timetable(subject_code, subject_name, semester_id, session_type)
    slot = TimetableSlot.query.filter_by(
        semester_id=semester_id,
        academic_year=academic_year,
        division=division,
        batch=batch,
        day_of_week=day,
        start_time=start,
        end_time=end,
        subject_code=subject_code,
    ).first()
    if not slot:
        slot = TimetableSlot(
            semester_id=semester_id,
            academic_year=academic_year,
            division=division,
            batch=batch,
            day_of_week=day,
            start_time=start,
            end_time=end,
            subject_code=subject_code,
        )
        db.session.add(slot)
    slot.faculty_name = faculty
    slot.session_type = session_type
    slot.room = room
    slot.is_active = True
    return slot


def _parse_visual_timetable_upload(wb, default_academic_year=None):
    ws = wb["Class Timetable"] if "Class Timetable" in wb.sheetnames else wb.active

    try:
        semester_id = int(ws["D4"].value)
    except (TypeError, ValueError):
        return False, ["Class Timetable: Semester value missing/invalid in cell D4."], 0

    academic_year = _clean(ws["B4"].value, default_academic_year or "2025-26")
    division = clean_filter_value(ws["H4"].value)

    errors = []
    created_or_updated = 0
    header_row = 6
    first_day_row = 7
    last_day_row = 12

    for col in range(2, ws.max_column + 1):
        start, end = _parse_time_range(ws.cell(header_row, col).value)
        if not start or not end:
            continue
        for row in range(first_day_row, last_day_row + 1):
            day = clean_day(ws.cell(row, 1).value)
            if not day:
                continue
            cell_value = ws.cell(row, col).value
            for entry_text in _split_visual_cell_entries(cell_value):
                parsed = _parse_visual_entry(entry_text)
                if not parsed:
                    continue
                subject_code = parsed["subject_code"]
                subject_name = parsed["subject_name"]
                batch = parsed["batch"]
                session_type = parsed["session_type"]
                faculty = parsed["faculty"]
                if not subject_code:
                    errors.append(f"Row {row}, column {col}: subject code required")
                    continue
                _upsert_timetable_slot(
                    semester_id=semester_id,
                    academic_year=academic_year,
                    division=division,
                    batch=batch,
                    day=day,
                    start=start,
                    end=end,
                    subject_code=subject_code,
                    subject_name=subject_name,
                    session_type=session_type,
                    faculty=faculty,
                    room=None,
                )
                created_or_updated += 1

    if errors:
        db.session.rollback()
        return False, errors[:20], 0
    if created_or_updated == 0:
        return False, ["No timetable subject entries found. Add subject codes inside the timetable cells."], 0

    db.session.commit()
    return True, [], created_or_updated

def _header_map(ws, row=4):
    mapping = {}
    for cell in ws[row]:
        if cell.value:
            key = str(cell.value).strip().lower().replace(" ", "_")
            mapping[key] = cell.column
    return mapping


def _parse_flat_timetable_upload(wb, default_academic_year=None):
    ws = wb["Timetable Upload"] if "Timetable Upload" in wb.sheetnames else wb.active
    headers = _header_map(ws, 4)
    required = ["semester", "day", "start_time", "end_time", "subject_code", "subject_name"]
    missing = [h for h in required if h not in headers]
    if missing:
        return False, [f"Missing required column(s): {', '.join(missing)}"], 0

    errors = []
    created_or_updated = 0

    for r in range(5, ws.max_row + 1):
        def val(name):
            col = headers.get(name)
            return ws.cell(r, col).value if col else None

        if all(val(h) in (None, "") for h in ["semester", "day", "start_time", "end_time", "subject_code"]):
            continue

        try:
            semester_id = int(val("semester"))
        except (TypeError, ValueError):
            errors.append(f"Row {r}: invalid semester")
            continue

        academic_year = _clean(val("academic_year"), default_academic_year or "2025-26")
        division = clean_filter_value(val("division"))
        batch = clean_filter_value(val("batch"))
        day = clean_day(val("day"))
        start = parse_time_value(val("start_time"))
        end = parse_time_value(val("end_time"))
        subject_code = clean_code(val("subject_code"))
        subject_name = _clean(val("subject_name"))
        session_type = clean_session_type(val("session_type"))
        faculty = _clean(val("faculty"), None)
        room = _clean(val("room"), None)

        if not day:
            errors.append(f"Row {r}: invalid day")
            continue
        if not start or not end:
            errors.append(f"Row {r}: invalid start/end time")
            continue
        if not subject_code or not subject_name:
            errors.append(f"Row {r}: subject code/name required")
            continue

        _upsert_timetable_slot(
            semester_id=semester_id,
            academic_year=academic_year,
            division=division,
            batch=batch,
            day=day,
            start=start,
            end=end,
            subject_code=subject_code,
            subject_name=subject_name,
            session_type=session_type,
            faculty=faculty,
            room=room,
        )
        created_or_updated += 1

    if errors:
        db.session.rollback()
        return False, errors[:20], 0

    db.session.commit()
    return True, [], created_or_updated


def parse_timetable_upload(file_obj, default_academic_year=None):
    wb = load_workbook(file_obj, data_only=True)

    # New teacher-friendly format, same visual idea as the provided class timetable.
    if "Class Timetable" in wb.sheetnames:
        return _parse_visual_timetable_upload(wb, default_academic_year)

    # Backward compatibility for the previous row-wise timetable template.
    return _parse_flat_timetable_upload(wb, default_academic_year)

def timetable_slots_for_subject(semester_id, academic_year, subject_code, division='ALL', batch='ALL'):
    division = clean_filter_value(division)
    batch = clean_filter_value(batch)
    q = TimetableSlot.query.filter_by(
        semester_id=semester_id,
        academic_year=academic_year,
        subject_code=subject_code,
        is_active=True,
    )
    # Include whole-class slots and selected division/batch slots.
    q = q.filter(TimetableSlot.division.in_(['ALL', division]))
    q = q.filter(TimetableSlot.batch.in_(['ALL', batch]))
    return q.order_by(TimetableSlot.day_of_week, TimetableSlot.start_time).all()


def generate_dates_from_slots(slots, start_date, end_date):
    if not slots or not start_date or not end_date or start_date > end_date:
        return []
    by_day = {}
    for slot in slots:
        by_day.setdefault(slot.day_of_week, []).append(slot)

    results = []
    cur = start_date
    while cur <= end_date:
        day = DAY_NAMES[cur.weekday()]
        for slot in by_day.get(day, []):
            results.append({
                'date': cur,
                'slot': slot,
                'label': f"{cur.strftime('%d-%b')}\n{slot.start_time.strftime('%H:%M')}-{slot.end_time.strftime('%H:%M')}",
            })
        cur += timedelta(days=1)
    results.sort(key=lambda x: (x['date'], x['slot'].start_time))
    return results


def students_for_attendance(subject_code, semester_id, academic_year, division='ALL', batch='ALL'):
    subject = Subject.query.get(subject_code)
    division = clean_filter_value(division)
    batch = clean_filter_value(batch)

    # Normal subjects use enrollment table.
    q = (
        db.session.query(Student)
        .join(Enrollment, Enrollment.prn == Student.prn)
        .filter(
            Enrollment.subject_code == subject_code,
            Enrollment.semester_id == semester_id,
            Enrollment.academic_year == academic_year,
        )
    )
    q = apply_student_batch_filters(q, Student, '' if division == 'ALL' else division, '' if batch == 'ALL' else batch)
    enrolled = q.order_by(Student.prn).all()
    if enrolled:
        return enrolled

    # Attendance-only subjects often do not have enrollment rows.
    q = Student.query
    q = apply_student_batch_filters(q, Student, '' if division == 'ALL' else division, '' if batch == 'ALL' else batch)
    return q.order_by(Student.prn).all()


def generate_attendance_template(meta, students, generated_sessions):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    title = f"Attendance Sheet - {meta.get('subject_code')}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(8, 4 + len(generated_sessions)))
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(name=TNR, size=14, bold=True, color=WHITE)
    ws.cell(1, 1).fill = PatternFill('solid', fgColor=DARK)
    ws.cell(1, 1).alignment = _center()

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(8, 4 + len(generated_sessions)))
    ws.cell(2, 1).value = (
        f"Semester {meta.get('semester_id')} | {meta.get('academic_year')} | "
        f"{meta.get('subject_name')} | {meta.get('division') or 'ALL'} / {meta.get('batch') or 'ALL'} | "
        f"Fill P/A/L only"
    )
    ws.cell(2, 1).font = Font(name=TNR, size=10, italic=True)
    ws.cell(2, 1).alignment = _center()

    headers = ["Sr", "PRN", "Student Name"]
    for item in generated_sessions:
        headers.append(item['label'])
    headers += ["Total", "Present", "%"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.font = Font(name=TNR, size=9, bold=True, color=WHITE)
        cell.fill = PatternFill('solid', fgColor=GOLD if col <= 3 else DARK)
        cell.alignment = _center()
        cell.border = _border()

    start_date_col = 4
    total_col = start_date_col + len(generated_sessions)
    present_col = total_col + 1
    pct_col = total_col + 2

    for r, stu in enumerate(students, 5):
        values = [r - 4, stu.prn, stu.name]
        for c, val in enumerate(values, 1):
            cell = ws.cell(r, c, val)
            cell.font = Font(name=TNR, size=10)
            cell.alignment = _center() if c != 3 else _left()
            cell.border = _border()
        for c in range(start_date_col, total_col):
            cell = ws.cell(r, c, "")
            cell.font = Font(name=TNR, size=10)
            cell.alignment = _center()
            cell.border = _border()
        row_range = f"{get_column_letter(start_date_col)}{r}:{get_column_letter(total_col-1)}{r}" if generated_sessions else ""
        ws.cell(r, total_col, f'=COUNTIF({row_range},"P")+COUNTIF({row_range},"A")' if generated_sessions else 0)
        ws.cell(r, present_col, f'=COUNTIF({row_range},"P")' if generated_sessions else 0)
        ws.cell(r, pct_col, f'=IF({get_column_letter(total_col)}{r}=0,"",ROUND({get_column_letter(present_col)}{r}/{get_column_letter(total_col)}{r}*100,2))' if generated_sessions else "")
        for c in (total_col, present_col, pct_col):
            ws.cell(r, c).font = Font(name=TNR, size=10, bold=True)
            ws.cell(r, c).alignment = _center()
            ws.cell(r, c).border = _border()

    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 32
    for c in range(start_date_col, total_col):
        ws.column_dimensions[get_column_letter(c)].width = 14
    for c in (total_col, present_col, pct_col):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.freeze_panes = "D5"
    ws.row_dimensions[4].height = 40

    # Hidden metadata for safe upload.
    meta_ws = wb.create_sheet("_meta")
    meta_ws.sheet_state = "hidden"
    rows = [
        ("semester_id", meta.get('semester_id')),
        ("academic_year", meta.get('academic_year')),
        ("subject_code", meta.get('subject_code')),
        ("division", meta.get('division') or 'ALL'),
        ("batch", meta.get('batch') or 'ALL'),
    ]
    for i, (k, v) in enumerate(rows, 1):
        meta_ws.cell(i, 1, k)
        meta_ws.cell(i, 2, v)
    meta_ws.cell(10, 1, "date")
    meta_ws.cell(10, 2, "start_time")
    meta_ws.cell(10, 3, "end_time")
    meta_ws.cell(10, 4, "session_type")
    for i, item in enumerate(generated_sessions, 11):
        slot = item['slot']
        meta_ws.cell(i, 1, item['date'].isoformat())
        meta_ws.cell(i, 2, slot.start_time.strftime('%H:%M'))
        meta_ws.cell(i, 3, slot.end_time.strftime('%H:%M'))
        meta_ws.cell(i, 4, slot.session_type)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", str(meta.get('subject_code') or 'subject'))
    return buf, f"Attendance_{safe}_Sem{meta.get('semester_id')}.xlsx"


def parse_attendance_upload(file_obj):
    wb = load_workbook(file_obj, data_only=False)
    if "Attendance" not in wb.sheetnames or "_meta" not in wb.sheetnames:
        return False, ["Invalid attendance file. Please upload the generated attendance template."], 0
    ws = wb["Attendance"]
    meta_ws = wb["_meta"]

    meta = {}
    for r in range(1, 8):
        k = meta_ws.cell(r, 1).value
        v = meta_ws.cell(r, 2).value
        if k:
            meta[str(k)] = v

    semester_id = int(meta.get('semester_id'))
    academic_year = str(meta.get('academic_year'))
    subject_code = clean_code(meta.get('subject_code'))
    division = clean_filter_value(meta.get('division'))
    batch = clean_filter_value(meta.get('batch'))

    sessions = []
    r = 11
    while meta_ws.cell(r, 1).value:
        lecture_date = parse_date(meta_ws.cell(r, 1).value)
        start = parse_time_value(meta_ws.cell(r, 2).value)
        end = parse_time_value(meta_ws.cell(r, 3).value)
        session_type = clean_session_type(meta_ws.cell(r, 4).value)
        if lecture_date and start and end:
            sess = AttendanceSession.query.filter_by(
                semester_id=semester_id,
                academic_year=academic_year,
                division=division,
                batch=batch,
                subject_code=subject_code,
                lecture_date=lecture_date,
                start_time=start,
                end_time=end,
            ).first()
            if not sess:
                sess = AttendanceSession(
                    semester_id=semester_id,
                    academic_year=academic_year,
                    division=division,
                    batch=batch,
                    subject_code=subject_code,
                    lecture_date=lecture_date,
                    start_time=start,
                    end_time=end,
                    session_type=session_type,
                    status='COMPLETED',
                )
                db.session.add(sess)
                db.session.flush()
            sessions.append(sess)
        r += 1

    if not sessions:
        return False, ["No attendance dates found in hidden metadata."], 0

    date_start_col = 4
    errors = []
    saved_entries = 0

    for row in range(5, ws.max_row + 1):
        prn = _clean(ws.cell(row, 2).value)
        if not prn:
            continue
        if not Student.query.get(prn):
            errors.append(f"Row {row}: PRN {prn} not found")
            continue
        attended = 0
        total = 0
        for i, sess in enumerate(sessions):
            raw = _clean(ws.cell(row, date_start_col + i).value).upper()
            if raw == "":
                continue
            if raw in PRESENT_VALUES:
                status = "P"
                total += 1
                attended += 1
            elif raw in ABSENT_VALUES:
                status = "A"
                total += 1
            elif raw in EXCUSED_VALUES:
                status = "L"
            else:
                errors.append(f"Row {row}: invalid attendance value '{raw}'")
                continue

            entry = AttendanceEntry.query.filter_by(session_id=sess.id, prn=prn).first()
            if not entry:
                entry = AttendanceEntry(session_id=sess.id, prn=prn)
                db.session.add(entry)
            entry.status = status
            saved_entries += 1

        summary = Attendance.query.filter_by(
            prn=prn,
            subject_code=subject_code,
            semester_id=semester_id,
            academic_year=academic_year,
        ).first()
        if not summary:
            summary = Attendance(
                prn=prn,
                subject_code=subject_code,
                semester_id=semester_id,
                academic_year=academic_year,
            )
            db.session.add(summary)
        summary.total_lectures = total
        summary.attended_lectures = attended
        summary.percentage = Decimal(str(round((attended / total * 100) if total else 0, 2)))

    if errors:
        db.session.rollback()
        return False, errors[:20], 0
    db.session.commit()
    return True, [], saved_entries


def attendance_summary(subject_code, semester_id, academic_year):
    rows = (
        db.session.query(Attendance, Student)
        .join(Student, Student.prn == Attendance.prn)
        .filter(
            Attendance.subject_code == subject_code,
            Attendance.semester_id == semester_id,
            Attendance.academic_year == academic_year,
        )
        .order_by(Student.prn)
        .all()
    )
    return rows


def generate_defaulter_excel(meta, rows, threshold=75.0):
    wb = Workbook()
    ws = wb.active
    ws.title = "Defaulters"
    headers = ["Sr", "PRN", "Student Name", "Total", "Present", "Percentage", "Status"]
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Attendance Defaulter List - {meta.get('subject_code')}"
    ws["A1"].font = Font(name=TNR, size=14, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill('solid', fgColor=DARK)
    ws["A1"].alignment = _center()
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.font = Font(name=TNR, bold=True, color=WHITE)
        cell.fill = PatternFill('solid', fgColor=GOLD)
        cell.alignment = _center()
        cell.border = _border()
    out_row = 4
    sr = 1
    for att, stu in rows:
        pct = float(att.percentage or 0)
        if pct >= threshold:
            continue
        status = "Critical" if pct < 65 else "Warning"
        vals = [sr, stu.prn, stu.name, att.total_lectures, att.attended_lectures, pct, status]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(out_row, c, v)
            cell.font = Font(name=TNR, size=10)
            cell.alignment = _center() if c != 3 else _left()
            cell.border = _border()
            if c == 7:
                cell.fill = PatternFill('solid', fgColor=RED if status == "Critical" else YELLOW)
        out_row += 1
        sr += 1
    for c, w in enumerate([7, 16, 32, 12, 12, 14, 14], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f"Attendance_Defaulters_{meta.get('subject_code')}.xlsx"


def generate_defaulter_pdf(meta, rows, threshold=75.0):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Title'], alignment=TA_CENTER, fontName='Times-Bold', fontSize=14)
    story = [
        Paragraph("CSMSS Chh. Shahu College of Engineering", title_style),
        Paragraph(f"Attendance Defaulter List - {meta.get('subject_code')} | Semester {meta.get('semester_id')} | {meta.get('academic_year')}", title_style),
        Spacer(1, 10),
    ]
    data = [["Sr", "PRN", "Student Name", "Total", "Present", "%", "Status"]]
    sr = 1
    for att, stu in rows:
        pct = float(att.percentage or 0)
        if pct >= threshold:
            continue
        data.append([sr, stu.prn, stu.name, att.total_lectures, att.attended_lectures, f"{pct:.2f}", "Critical" if pct < 65 else "Warning"])
        sr += 1
    if len(data) == 1:
        data.append(["-", "-", "No defaulters below threshold", "-", "-", "-", "-"])
    table = Table(data, repeatRows=1, colWidths=[1.1*cm, 3*cm, 8*cm, 2*cm, 2*cm, 2*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1A2B4A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Times-Bold'),
        ('FONTNAME', (0,1), (-1,-1), 'Times-Roman'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (2,1), (2,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(table)
    doc.build(story)
    buf.seek(0)
    return buf, f"Attendance_Defaulters_{meta.get('subject_code')}.pdf"
