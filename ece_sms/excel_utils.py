import io
import statistics
import re
import os
from openpyxl.drawing.image import Image
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from models import db, Student, TheoryMarks, Enrollment
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

LEFT_LOGO = os.path.join(BASE_DIR, "static", "images", "left_logo.jpg")
RIGHT_LOGO = os.path.join(BASE_DIR, "static", "images", "right_logo.png")

TNR   = "Times New Roman"
BLACK = "000000"
WHITE = "FFFFFF"
LGRAY = "F2F2F2"
MGRAY = "D0D0D0"
DGRAY = "888888"
HGRAY = "808080"   # heading gray — replaces black


def _thin_border():
    s = Side(style='thin', color=BLACK)
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_border():
    s = Side(style='medium', color=BLACK)
    return Border(left=s, right=s, top=s, bottom=s)

def _body_font(size=10, bold=False):
    return Font(name=TNR, size=size, bold=bold, color=BLACK)

def _hdr_font(size=10):
    return Font(name=TNR, size=size, bold=True, color=WHITE)

def _center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def _left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def _right():
    return Alignment(horizontal='right', vertical='center', wrap_text=True)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _sanitize(name):
    name = name.strip()
    name = re.sub(r'[^\w\s-]', '', name)
    name = re.sub(r'[\s]+', '_', name)
    return name

def _roll(sr):
    """Generate roll number EC3101 ... EC3175"""
    return f"EC31{sr:02d}"


# ══════════════════════════════════════════════════════════════
#  1. BLANK TEMPLATE
# ══════════════════════════════════════════════════════════════
def generate_ct1_template(meta, students):
    import os
    from openpyxl.drawing.image import Image

    wb = Workbook()
    ws = wb.active
    ws.title = f"{meta.get('exam_type','CT1')} Marks"

    # COLUMN WIDTHS
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 42
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    exam_label = meta.get('exam_label', meta.get('exam_type', 'CT1'))
    max_marks = meta.get('max_marks', 10)

    # ---------------- HEADER ----------------

    for i in range(1, 6):
        ws.row_dimensions[i].height = 24

    ws.row_dimensions[1].height = 70

    # LEFT LOGO
    if os.path.exists(LEFT_LOGO):
        left_logo = Image(LEFT_LOGO)
        left_logo.width = 95
        left_logo.height = 95
        ws.add_image(left_logo, "A1")

    # RIGHT LOGO
    if os.path.exists(RIGHT_LOGO):
        right_logo = Image(RIGHT_LOGO)
        right_logo.width = 95
        right_logo.height = 95
        ws.add_image(right_logo, "F1")

    # CENTER HEADER
    ws.merge_cells("B1:E1")
    ws["B1"] = "Chhatrapati Shahu Maharaj Shikshan Sanstha"

    ws.merge_cells("B2:E2")
    ws["B2"] = "CSMSS Chh. Shahu College of Engineering"

    ws.merge_cells("B3:E3")
    ws["B3"] = "Kanchanwadi, Chhatrapati Sambhajinagar – 431011 (Maharashtra)"

    ws.merge_cells("B4:E5")
    ws["B4"] = "Department of Electronics and Computer Engineering"

    for r in range(1, 6):
        c = ws[f'B{r}']
        c.alignment = _center()
        c.border = _thick_border()

    ws["B1"].font = Font(name=TNR, size=12, bold=False)
    ws["B2"].font = Font(name=TNR, size=20, bold=True)
    ws["B3"].font = Font(name=TNR, size=11)
    ws["B4"].font = Font(name=TNR, size=18, bold=True)

    # BORDER A1:F5
    for row in range(1, 6):
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = _thick_border()

    # ---------------- DETAILS ----------------

    row = 6

    ws[f"A{row}"] = "Subject Code :"
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"D{row}"] = "Subject Name :"
    ws.merge_cells(f"E{row}:F{row}")
    ws[f"B{row}"] = meta['subject_code']
    ws[f"E{row}"] = meta['subject_name']

    row = 7

    ws[f"A{row}"] = "Exam Type :"
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"D{row}"] = "Subject Teacher :"
    ws.merge_cells(f"E{row}:F{row}")
    ws[f"B{row}"] = exam_label
    ws[f"E{row}"] = meta['subject_teacher']

    row = 8

    ws[f"A{row}"] = "Exam Date :"
    ws.merge_cells(f"B{row}:F{row}")
    ws[f"B{row}"] = meta['exam_date']

    for r in range(6, 9):
        for c in range(1, 7):
            ws.cell(r, c).border = _thick_border()
            ws.cell(r, c).font = _body_font(12, bold=(c in [1,4]))
            ws.cell(r, c).alignment = _left()

    # ---------------- TABLE HEADER ----------------

    header_row = 10

    headers = [
        "Sr No",
        "Roll No",
        "PRN No",
        "Student Name",
        f"Marks\n(Out of {max_marks})",
        ""
    ]

    for col, hdr in enumerate(headers, 1):
        c = ws.cell(header_row, col, hdr)
        c.font = Font(name=TNR, size=13, bold=True)
        c.alignment = _center()
        c.border = _thick_border()

    ws.merge_cells(f"E{header_row}:F{header_row}")
    ws.row_dimensions[header_row].height = 38

    # ---------------- STUDENT ROWS ----------------

    for sr, student in enumerate(students, 1):
        r = header_row + sr

        ws.cell(r,1,sr)
        ws.cell(r,2,_roll(sr))
        ws.cell(r,3,student.prn)
        ws.cell(r,4,student.name)

        ws.merge_cells(f"E{r}:F{r}")

        for c in range(1,7):
            ws.cell(r,c).border = _thin_border()
            ws.cell(r,c).font = _body_font(12)
            ws.cell(r,c).alignment = _center()

        ws.cell(r,4).alignment = _left()
        ws.row_dimensions[r].height = 24

    # ---------------- NOTE ----------------

    note_row = header_row + len(students) + 1
    ws.merge_cells(f"A{note_row}:F{note_row}")

    ws[f"A{note_row}"] = (
        f"NOTE: Enter marks in last column (0 to {max_marks}). "
        "Do not edit Sr No, Roll No, PRN No or Student Name."
    )

    ws[f"A{note_row}"].font = Font(name=TNR, size=12, bold=True)
    ws[f"A{note_row}"].alignment = _left()
    ws[f"A{note_row}"].border = _thick_border()

    # ---------------- SAVE ----------------

    exam_type = meta.get('exam_type', 'CT1')
    subj_clean = _sanitize(meta['subject_name'])
    filename = f"{subj_clean}_{exam_type}_Template.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return buf, filename
# ══════════════════════════════════════════════════════════════
#  2. PARSE UPLOAD  (unchanged logic)
# ══════════════════════════════════════════════════════════════
def parse_ct1_upload(file_obj, subject_code, semester_id,
                     academic_year, exam_type='CT1', max_marks=10):
    try:
        wb = load_workbook(file_obj, data_only=True)
        ws = wb.active
    except Exception as e:
        return False, [{'row': 0, 'col': 'FILE',
                        'msg': f'Cannot read file: {str(e)}'}], 0

    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and 'Roll No' in str(cell.value):
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        return False, [{'row': 0, 'col': 'STRUCTURE',
                        'msg': 'Header row not found. Use the official template.'}], 0

    valid_prns    = {s.prn for s in Student.query.all()}
    ABSENT_VALUES = {'ab', 'absent', 'a', ''}
    errors        = []
    records       = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row[2]:   # PRN is now col C (index 2)
            continue
        prn_val   = str(row[2]).strip()
        marks_val = row[4]   # Marks now col E (index 4)

        if prn_val.lower().startswith('note'):
            continue

        row_num = header_row + 1 + len(records) + len(errors)

        if prn_val not in valid_prns:
            errors.append({'row': row_num, 'col': 'PRN',
                           'msg': f'PRN {prn_val} not found in database'})
            continue

        if marks_val is None:
            records.append({'prn': prn_val, 'ct1': None, 'absent': True})
            continue

        str_val = str(marks_val).strip().lower()
        if str_val in ABSENT_VALUES:
            records.append({'prn': prn_val, 'ct1': None, 'absent': True})
            continue

        try:
            marks = float(str_val)
        except (ValueError, TypeError):
            errors.append({'row': row_num, 'col': 'Marks',
                           'msg': f'Value "{marks_val}" is not a number or valid absent marker.'})
            continue

        if not (0 <= marks <= max_marks):
            errors.append({'row': row_num, 'col': 'Marks',
                           'msg': f'Marks {marks} out of range. Must be 0 to {max_marks}.'})
            continue

        records.append({'prn': prn_val, 'ct1': marks, 'absent': False})

    if errors:
        return False, errors, 0

    field_map = {'CT1': 'ct1', 'CT2': 'ct2',
                 'ASSIGNMENT': 'assignment', 'MIDSEM': 'midsem'}
    field = field_map.get(exam_type.upper(), 'ct1')

    from calculations import update_internal_totals

    count = 0
    for rec in records:
        existing = TheoryMarks.query.filter_by(
            prn=rec['prn'], subject_code=subject_code,
            academic_year=academic_year
        ).first()
        if existing:
            setattr(existing, field, rec['ct1'])
            update_internal_totals(existing)
        else:
            kwargs = {
                'prn': rec['prn'], 'subject_code': subject_code,
                'semester_id': semester_id, 'academic_year': academic_year,
                field: rec['ct1']
            }
            new_row = TheoryMarks(**kwargs)
            update_internal_totals(new_row)
            db.session.add(new_row)
        count += 1

    db.session.commit()
    return True, [], count


# ══════════════════════════════════════════════════════════════
#  3. EXCEL REPORT
# ══════════════════════════════════════════════════════════════
def generate_ct1_excel_report(meta, data):
    import io
    import os
    import statistics
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = f"{meta.get('exam_type','CT1')} Report"

    # COLUMN WIDTHS
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 42
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    exam_label = meta.get('exam_label', meta.get('exam_type', 'CT1'))
    max_marks = meta.get('max_marks', 10)
    pass_mark = max_marks * 0.4

    # ==========================================================
    # HEADER
    # ==========================================================

    for i in range(1, 8):
        ws.row_dimensions[i].height = 28

    # Left Logo
    if os.path.exists(LEFT_LOGO):
        img = Image(LEFT_LOGO)
        img.width = 95
        img.height = 95
        ws.add_image(img, "A1")

    if os.path.exists(RIGHT_LOGO):
        img2 = Image(RIGHT_LOGO)
        img2.width = 95
        img2.height = 95
        ws.add_image(img2, "F1")

    ws.merge_cells("B1:E1")
    ws["B1"] = "Chhatrapati Shahu Maharaj Shikshan Sanstha, Aurangabad"

    ws.merge_cells("B2:E2")
    ws["B2"] = "CSMSS Chh. Shahu College of Engineering"

    ws.merge_cells("B3:E3")
    ws["B3"] = "Kanchanwadi, Chhatrapati Sambhajinagar – 431011"

    ws.merge_cells("B4:E5")
    ws["B4"] = "Department of Electronics and Computer Engineering"

    ws["B1"].font = Font(name=TNR, size=12)
    ws["B2"].font = Font(name=TNR, size=20, bold=True)
    ws["B3"].font = Font(name=TNR, size=11)
    ws["B4"].font = Font(name=TNR, size=16, bold=True)

    for r in range(1, 6):
        for c in range(1, 7):
            ws.cell(r, c).border = _thick_border()

    for c in ["B1", "B2", "B3", "B4"]:
        ws[c].alignment = _center()

    # ==========================================================
    # DETAILS
    # ==========================================================

    row = 8

    ws[f"A{row}"] = "Subject Code :"
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"] = meta['subject_code']

    ws[f"D{row}"] = "Subject Name :"
    ws.merge_cells(f"E{row}:F{row}")
    ws[f"E{row}"] = meta['subject_name']

    row = 9

    ws[f"A{row}"] = "Exam Type :"
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"B{row}"] = exam_label

    ws[f"D{row}"] = "Subject Teacher :"
    ws.merge_cells(f"E{row}:F{row}")
    ws[f"E{row}"] = meta['subject_teacher']

    row = 10

    ws[f"A{row}"] = "Exam Date :"
    ws.merge_cells(f"B{row}:F{row}")
    ws[f"B{row}"] = meta['exam_date']

    for r in range(8, 11):
        for c in range(1, 7):
            ws.cell(r, c).border = _thick_border()
            ws.cell(r, c).font = _body_font(12)
            ws.cell(r, c).alignment = _left()

    # ==========================================================
    # TABLE HEADER
    # ==========================================================

    hr = 12

    headers = [
        "Sr No",
        "Roll No",
        "PRN No",
        "Student Name",
        f"Marks Obtained\n(Out of {max_marks})",
        ""
    ]

    for col, hdr in enumerate(headers, 1):
        c = ws.cell(hr, col, hdr)
        c.font = Font(name=TNR, size=13, bold=True)
        c.alignment = _center()
        c.border = _thick_border()

    ws.merge_cells(f"E{hr}:F{hr}")
    ws.row_dimensions[hr].height = 36

    # ==========================================================
    # DATA
    # ==========================================================

    marks_list = []

    for d in data:
        r = hr + d['sr']
        is_absent = d['ct1'] is None

        vals = [
            d['sr'],
            _roll(d['sr']),
            d['prn'],
            d['name'],
            "AB" if is_absent else d['ct1']
        ]

        for c_idx, val in enumerate(vals, 1):
            c = ws.cell(r, c_idx, val)
            c.font = _body_font(12, bold=is_absent)
            c.alignment = _center() if c_idx != 4 else _left()
            c.border = _thin_border()

        ws.merge_cells(f"E{r}:F{r}")
        ws.row_dimensions[r].height = 24

        if not is_absent:
            marks_list.append(d['ct1'])

    # ==========================================================
    # ANALYSIS
    # ==========================================================

    total_students = len(data)
    present_students = len(marks_list)
    absent_students = total_students - present_students
    pass_students = sum(1 for m in marks_list if m >= pass_mark)
    fail_students = present_students - pass_students

    pass_pct = round(pass_students / present_students * 100, 2) if present_students else 0
    fail_pct = round(fail_students / present_students * 100, 2) if present_students else 0
    avg = round(sum(marks_list) / len(marks_list), 2) if marks_list else 0
    med = round(statistics.median(marks_list), 2) if marks_list else 0
    high_val = max(marks_list) if marks_list else 0
    low_val = min(marks_list) if marks_list else 0

    as_start = hr + len(data) + 2

    analysis_rows = [
        ["Total", total_students, "Present", present_students, "Absent", absent_students],
        ["Pass", pass_students, "Fail", fail_students, "Pass %", f"{pass_pct}%"],
        ["Average", avg, "Median", med, "Highest", high_val],
        ["Lowest", low_val, "Fail %", f"{fail_pct}%", "", ""]
    ]

    for idx, vals in enumerate(analysis_rows):
        rr = as_start + idx
        for col, val in enumerate(vals, 1):
            c = ws.cell(rr, col, val)
            c.font = _body_font(11, bold=True)
            c.alignment = _center()
            c.border = _thick_border()

    # ==========================================================
    # NOTE
    # ==========================================================

    note_row = as_start + 5
    ws.merge_cells(f"A{note_row}:F{note_row}")

    c = ws[f"A{note_row}"]
    c.value = "Report generated by ECE Student Management System"
    c.font = Font(name=TNR, size=11, italic=True)
    c.alignment = _center()
    c.border = _thick_border()

    # ==========================================================
    # SIGNATURES
    # ==========================================================

    sig_row = note_row + 2
    ws.row_dimensions[sig_row].height = 80

    ws.merge_cells(f"A{sig_row}:B{sig_row}")
    ws.merge_cells(f"C{sig_row}:D{sig_row}")
    ws.merge_cells(f"E{sig_row}:F{sig_row}")

    ws[f"A{sig_row}"] = "Subject Teacher\n(Signature)"
    ws[f"C{sig_row}"] = "Prof. A. V. Khake\nClass Teacher\n(Signature)"
    ws[f"E{sig_row}"] = "Dr. D. L. Bhuyar\nHOD\n(Signature)"

    for cell in [f"A{sig_row}", f"C{sig_row}", f"E{sig_row}"]:
        ws[cell].font = Font(name=TNR, size=13, bold=True)
        ws[cell].alignment = Alignment(
            horizontal='center',
            vertical='bottom',
            wrap_text=True
        )
        ws[cell].border = _thick_border()

    # SAVE
    exam_type = meta.get('exam_type', 'CT1')
    subj_clean = _sanitize(meta['subject_name'])
    filename = f"{subj_clean}_{exam_type}_Report.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return buf, filename

# ══════════════════════════════════════════════════════════════
#  INTERNAL MARKS REPORT — EXCEL
# ══════════════════════════════════════════════════════════════
def generate_internal_excel(meta, data):
    from openpyxl.drawing.image import Image

    wb = Workbook()
    ws = wb.active
    ws.title = "Internal Marks"

    # ===== LOGOS =====
    left_logo = Image(LEFT_LOGO)
    right_logo = Image(RIGHT_LOGO)

    left_logo.width = 85
    left_logo.height = 85
    right_logo.width = 85
    right_logo.height = 85

    ws.add_image(left_logo, "A1")
    ws.add_image(right_logo, "I1")

    last_col = 9
    last_letter = get_column_letter(last_col)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 28
    for col in range(4, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13

    # ===== HEADER =====
    ws.merge_cells('B1:H1')
    c = ws['B1']
    c.value = (
        "Chhatrapati Shahu Maharaj Shikshan Sanstha's\n"
        "CSMSS Chh. Shahu College of Engineering\n"
        "Kanchanwadi, Chhatrapati Sambhajinagar\n"
        "Department of Electronics and Computer Engineering"
    )
    c.font = Font(name=TNR, size=13, bold=True)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = _thin_border()
    ws.row_dimensions[1].height = 95

    ws.row_dimensions[2].height = 6

    # ===== SUBJECT INFO =====
    ws.merge_cells('A3:B3')
    ws['A3'] = "Subject Code :"
    ws['A3'].font = Font(name=TNR, size=10, bold=True)
    ws['A3'].border = _thin_border()

    ws.merge_cells('C3:D3')
    ws['C3'] = meta['subject_code']
    ws['C3'].border = _thin_border()

    ws.merge_cells('E3:F3')
    ws['E3'] = "Subject Name :"
    ws['E3'].font = Font(name=TNR, size=10, bold=True)
    ws['E3'].border = _thin_border()

    ws.merge_cells('G3:I3')
    ws['G3'] = meta['subject_name']
    ws['G3'].border = _thin_border()

    ws.merge_cells('A4:B4')
    ws['A4'] = "Semester :"
    ws['A4'].font = Font(name=TNR, size=10, bold=True)
    ws['A4'].border = _thin_border()

    ws.merge_cells('C4:D4')
    ws['C4'] = meta['semester_id']
    ws['C4'].border = _thin_border()

    ws.merge_cells('E4:F4')
    ws['E4'] = "Academic Year :"
    ws['E4'].font = Font(name=TNR, size=10, bold=True)
    ws['E4'].border = _thin_border()

    ws.merge_cells('G4:I4')
    ws['G4'] = meta['academic_year']
    ws['G4'].border = _thin_border()

    # ===== TABLE HEADER =====
    hr = 6
    headers = [
        'SR', 'PRN', 'NAME',
        'CT1', 'CT2',
        'ASSIGNMENT', 'CA',
        'MIDSEM', 'INTERNAL'
    ]

    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=hr, column=col, value=hdr)
        c.font = Font(name=TNR, size=10, bold=True)
        c.alignment = _center()
        c.border = _thin_border()

    # ===== DATA =====
    for d in data:
        row_idx = hr + d['sr']

        vals = [
            d['sr'],
            d['prn'],
            d['name'],
            d['ct1'],
            d['ct2'],
            d['assignment'],
            d['ca'],
            d['midsem'],
            d['internal']
        ]

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.font = _body_font()
            c.alignment = _center() if col != 3 else _left()
            c.border = _thin_border()

    # ===== SIGNATURES =====
    sig_row = hr + len(data) + 3
    ws.row_dimensions[sig_row].height = 60

    ws.merge_cells(f'A{sig_row}:C{sig_row}')
    ws.merge_cells(f'D{sig_row}:F{sig_row}')
    ws.merge_cells(f'G{sig_row}:I{sig_row}')

    ws[f'A{sig_row}'] = "Subject Teacher"
    ws[f'D{sig_row}'] = "Prof. A. V. Khake\nClass Teacher"
    ws[f'G{sig_row}'] = "Dr. D. L. Bhuyar\nHOD"

    for cell in [f'A{sig_row}', f'D{sig_row}', f'G{sig_row}']:
        ws[cell].font = Font(name=TNR, size=11, bold=True)
        ws[cell].alignment = Alignment(
            horizontal='center',
            vertical='bottom',
            wrap_text=True
        )
        ws[cell].border = _thick_border()

    # ===== FOOTER =====
    footer_row = sig_row + 2
    ws.merge_cells(f'A{footer_row}:I{footer_row}')
    ws[f'A{footer_row}'] = (
        "Generated by ECE Student Management System — "
        "CSMSS Chh. Shahu College of Engineering"
    )
    ws[f'A{footer_row}'].font = Font(name=TNR, size=9, italic=True)
    ws[f'A{footer_row}'].alignment = _center()

    subj_clean = _sanitize(meta['subject_name'])
    ay_clean = meta['academic_year'].replace('-', '_')

    filename = f"{subj_clean}_Internal_Sem{meta['semester_id']}_{ay_clean}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return buf, filename
# ══════════════════════════════════════════════════════════════
#  DEPARTMENT CT1 SUMMARY REPORT — EXCEL
#  Append after all existing functions in excel_utils.py
# ══════════════════════════════════════════════════════════════
def generate_ct1_summary_excel(summary_data):
    """
    Generates department-level CT1 summary Excel report
    formatted exactly like Internal Report style
    """

    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image

    meta = summary_data["report_meta"]
    cls_sum = summary_data["class_summary"]
    subjects = summary_data["subjects"]
    fac_table = summary_data["faculty_table"]
    sigs = summary_data["signatories"]

    n_subjects = len(subjects)
    total_cols = n_subjects + 2

    wb = Workbook()
    ws = wb.active
    ws.title = "CT1 Summary"

    # ==========================================================
    # LOGOS
    # ==========================================================
    if os.path.exists(LEFT_LOGO):
        left_logo = Image(LEFT_LOGO)
        left_logo.width = 85
        left_logo.height = 85
        ws.add_image(left_logo, "A1")

    if os.path.exists(RIGHT_LOGO):
        right_logo = Image(RIGHT_LOGO)
        right_logo.width = 85
        right_logo.height = 85
        ws.add_image(right_logo, f"{get_column_letter(total_cols)}1")

    last_col = total_cols
    last_letter = get_column_letter(last_col)

    # ==========================================================
    # COLUMN WIDTHS
    # ==========================================================
    ws.column_dimensions['A'].width = 26

    for i in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    # ==========================================================
    # HEADER
    # ==========================================================
    ws.merge_cells(f'B1:{get_column_letter(last_col-1)}1')

    c = ws['B1']
    c.value = (
        "Chhatrapati Shahu Maharaj Shikshan Sanstha's\n"
        "CSMSS Chh. Shahu College of Engineering\n"
        "Kanchanwadi, Chhatrapati Sambhajinagar\n"
        "Department of Electronics and Computer Engineering"
    )

    c.font = Font(name=TNR, size=13, bold=True)
    c.alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    c.border = _thin_border()

    ws.row_dimensions[1].height = 95
    ws.row_dimensions[2].height = 6

    # ==========================================================
    # REPORT TITLE
    # ==========================================================
    ws.merge_cells(f'A3:{last_letter}3')
    ws['A3'] = meta["report_title"]
    ws['A3'].font = Font(name=TNR, size=12, bold=True)
    ws['A3'].alignment = _center()
    ws['A3'].border = _thin_border()

    # ==========================================================
    # YEAR / DATE
    # ==========================================================
    ws.merge_cells('A4:C4')
    ws['A4'] = f"Academic Year : {meta['academic_year']}"
    ws['A4'].font = Font(name=TNR, size=10, bold=True)
    ws['A4'].border = _thin_border()

    ws.merge_cells(f'D4:{last_letter}4')
    ws['D4'] = f"Date : {meta['date_range']}"
    ws['D4'].alignment = _right()
    ws['D4'].border = _thin_border()

    # ==========================================================
    # CLASS
    # ==========================================================
    ws.merge_cells(f'A5:{last_letter}5')
    ws['A5'] = f"Class : {cls_sum['class_name']}    Division : {cls_sum['division']}"
    ws['A5'].font = Font(name=TNR, size=11, bold=True)
    ws['A5'].alignment = _center()
    ws['A5'].border = _thin_border()

    # ==========================================================
    # SUBJECT HEADER
    # ==========================================================
    hr = 7

    headers = ["Subject Name"] + [s["abbr"] for s in subjects]

    for col, hdr in enumerate(headers, 1):
        c = ws.cell(hr, col, hdr)
        c.font = Font(name=TNR, size=10, bold=True)
        c.alignment = _center()
        c.border = _thin_border()

    # ==========================================================
    # FULL SUBJECT NAMES
    # ==========================================================
    fnr = hr + 1

    ws.cell(fnr, 1, "Full Subject Name")

    for col in range(1, last_col + 1):
        ws.cell(fnr, col).border = _thin_border()
        ws.cell(fnr, col).alignment = _center()

    for idx, subj in enumerate(subjects):
        ws.cell(fnr, idx + 2, subj["name"])

    # ==========================================================
    # STATS
    # ==========================================================
    stats = [
        ("appeared", "Appeared Students"),
        ("passed", "Passed Students"),
        ("absent", "Absent"),
        ("failed", "Failed"),
        ("result_pct", "Subject Percentage Result")
    ]

    row = fnr + 1

    for key, label in stats:

        ws.cell(row, 1, label)

        for idx, subj in enumerate(subjects):
            val = subj[key]
            if key == "result_pct":
                val = f"{val}%"

            ws.cell(row, idx + 2, val)

        for col in range(1, last_col + 1):
            ws.cell(row, col).border = _thin_border()
            ws.cell(row, col).alignment = _center()

        row += 1

    # ==========================================================
    # AVERAGE RESULT
    # ==========================================================
    ws.merge_cells(f'B{row}:{last_letter}{row}')

    ws.cell(row, 1, "Average Class Result")
    ws.cell(row, 2, f"{cls_sum['average_class_result']}%")

    for col in range(1, last_col + 1):
        ws.cell(row, col).border = _thin_border()
        ws.cell(row, col).alignment = _center()

    # ==========================================================
    # FACULTY DETAILS
    # ==========================================================
    row += 2

    ws.merge_cells(f'A{row}:{last_letter}{row}')
    ws[f'A{row}'] = "Subject Faculty Details"
    ws[f'A{row}'].font = Font(name=TNR, bold=True)
    ws[f'A{row}'].alignment = _center()
    ws[f'A{row}'].border = _thin_border()

    row += 1

    headers = ["Abbreviation", "Subject Name", "Faculty"]

    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.font = Font(name=TNR, bold=True)
        c.alignment = _center()
        c.border = _thin_border()

    for f in fac_table:
        row += 1

        vals = [f["abbr"], f["name"], f["faculty"]]

        for i, v in enumerate(vals, 1):
            c = ws.cell(row, i, v)
            c.border = _thin_border()
            c.alignment = _left() if i != 1 else _center()

    # ==========================================================
    # SIGNATURES
    # ==========================================================
    row += 3
    ws.row_dimensions[row].height = 60

    sig_cells = [
        (1, sigs["coordinator"], "Class Test Coordinator"),
        (3, sigs["hod"], "HOD"),
        (5, sigs["dean"], "Dean Academics"),
        (7, sigs["principal"], "Principal"),
    ]

    for col, name, title in sig_cells:
        ws.merge_cells(
            start_row=row,
            start_column=col,
            end_row=row,
            end_column=min(col+1, last_col)
        )

        c = ws.cell(row, col)
        c.value = f"{name}\n{title}"
        c.font = Font(name=TNR, size=10, bold=True)
        c.alignment = Alignment(
            horizontal='center',
            vertical='bottom',
            wrap_text=True
        )
        c.border = _thick_border()

    # ==========================================================
    # FOOTER
    # ==========================================================
    row += 2

    ws.merge_cells(f'A{row}:{last_letter}{row}')
    ws[f'A{row}'] = (
        "Generated by ECE Student Management System — "
        "CSMSS Chh. Shahu College of Engineering"
    )
    ws[f'A{row}'].font = Font(name=TNR, size=9, italic=True)
    ws[f'A{row}'].alignment = _center()

    # ==========================================================
    # FREEZE
    # ==========================================================
    ws.freeze_panes = "B7"

    # ==========================================================
    # SAVE
    # ==========================================================
    exam_type = meta["exam_type"]
    ay = meta["academic_year"].replace("-", "_")
    class_label = cls_sum["class_name"].replace(" ", "")

    filename = f"{class_label}_{exam_type}_Summary_{ay}.xlsx"

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return buf, filename