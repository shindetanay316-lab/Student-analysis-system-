import io
import re
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from models import Student, Enrollment, TheoryMarks

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LEFT_LOGO = os.path.join(BASE_DIR, "static", "images", "left_logo.jpg")
RIGHT_LOGO = os.path.join(BASE_DIR, "static", "images", "right_logo.png")

TNR = "Times New Roman"
BLACK = "000000"
WHITE = "FFFFFF"

def _thin_border():
    s = Side(style='thin', color=BLACK)
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_border():
    s = Side(style='medium', color=BLACK)
    return Border(left=s, right=s, top=s, bottom=s)

def _body_font(size=10, bold=False):
    return Font(name=TNR, size=size, bold=bold, color=BLACK)

def _center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def _left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def _sanitize(name):
    name = name.strip()
    name = re.sub(r'[^\w\s-]', '', name)
    name = re.sub(r'[\s]+', '_', name)
    return name

def _roll(sr):
    return f"EC31{sr:02d}"

def generate_external_template(meta, students):
    wb = Workbook()
    ws = wb.active
    ws.title = "External Marks"

    # COLUMN WIDTHS
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 24

    # ---------------- HEADER ----------------
    for i in range(1, 6):
        ws.row_dimensions[i].height = 24
    ws.row_dimensions[1].height = 70

    # LEFT LOGO
    if os.path.exists(LEFT_LOGO):
        left_logo = Image(LEFT_LOGO)
        left_logo.width = 95
        left_logo.height = 95
        ws.add_image(left_logo, "A1")

    # RIGHT LOGO
    if os.path.exists(RIGHT_LOGO):
        right_logo = Image(RIGHT_LOGO)
        right_logo.width = 95
        right_logo.height = 95
        ws.add_image(right_logo, "E1")

    # CENTER HEADER
    ws.merge_cells("B1:D1")
    ws["B1"] = "Chhatrapati Shahu Maharaj Shikshan Sanstha"
    ws.merge_cells("B2:D2")
    ws["B2"] = "CSMSS Chh. Shahu College of Engineering"
    ws.merge_cells("B3:D3")
    ws["B3"] = "Kanchanwadi, Chhatrapati Sambhajinagar – 431011 (Maharashtra)"
    ws.merge_cells("B4:D5")
    ws["B4"] = "Department of Electronics and Computer Engineering"

    for r in range(1, 6):
        c = ws[f'B{r}']
        c.alignment = _center()
        c.border = _thick_border()

    ws["B1"].font = Font(name=TNR, size=12, bold=False)
    ws["B2"].font = Font(name=TNR, size=20, bold=True)
    ws["B3"].font = Font(name=TNR, size=11)
    ws["B4"].font = Font(name=TNR, size=18, bold=True)

    # BORDER A1:E5
    for row in range(1, 6):
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = _thick_border()

    # ---------------- DETAILS ----------------
    row = 6
    ws[f"A{row}"] = "Subject Code :"
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"D{row}"] = "Subject Name :"
    ws.cell(row=row, column=5)
    ws[f"B{row}"] = meta['subject_code']
    ws[f"E{row}"] = meta['subject_name']

    row = 7
    ws[f"A{row}"] = "Exam Type :"
    ws.merge_cells(f"B{row}:C{row}")
    ws[f"D{row}"] = "Subject Teacher :"
    ws.cell(row=row, column=5)
    ws[f"B{row}"] = "External Marks"
    ws[f"E{row}"] = meta.get('subject_teacher', 'N/A')

    row = 8
    ws[f"A{row}"] = "Exam Date :"
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"] = meta.get('exam_date', 'N/A') + (" | " + meta.get("batch_label", "All Students") if meta.get("batch_label") else "")

    for r in range(6, 9):
        for c in range(1, 6):
            ws.cell(r, c).border = _thick_border()
            ws.cell(r, c).font = _body_font(12, bold=(c in [1,4]))
            ws.cell(r, c).alignment = _left()

    # ---------------- TABLE HEADER ----------------
    header_row = 10
    headers = [
        "SR",
        "Roll No",
        "PRN",
        "Student Name",
        "External Marks /60"
    ]

    for col, hdr in enumerate(headers, 1):
        c = ws.cell(header_row, col, hdr)
        c.font = Font(name=TNR, size=13, bold=True)
        c.alignment = _center()
        c.border = _thick_border()

    ws.row_dimensions[header_row].height = 38

    # ---------------- STUDENT ROWS ----------------
    for sr, student in enumerate(students, 1):
        r = header_row + sr
        ws.cell(r, 1, sr)
        ws.cell(r, 2, _roll(sr))
        ws.cell(r, 3, student['prn'])
        ws.cell(r, 4, student['name'])
        ws.cell(r, 5, student.get('external_marks'))

        for c in range(1, 6):
            ws.cell(r, c).border = _thin_border()
            ws.cell(r, c).font = _body_font(12)
            ws.cell(r, c).alignment = _center()

        ws.cell(r, 4).alignment = _left()
        ws.row_dimensions[r].height = 24

        # Read-only fill style for metadata columns
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="F2F3F4")
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="F2F3F4")
        ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor="F2F3F4")
        ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor="F2F3F4")

    # ---------------- NOTE ----------------
    note_row = header_row + len(students) + 1
    ws.merge_cells(f"A{note_row}:E{note_row}")
    ws[f"A{note_row}"] = (
        "NOTE: Enter marks in the last column (0 to 60). "
        "Do not edit SR, Roll No, PRN, or Student Name."
    )
    ws[f"A{note_row}"].font = Font(name=TNR, size=12, bold=True)
    ws[f"A{note_row}"].alignment = _left()
    ws[f"A{note_row}"].border = _thick_border()

    # ---------------- PROTECTION ----------------
    ws.protection.sheet = True
    ws.protection.password = "ecesms"
    for i in range(1, len(students) + 1):
        ws.cell(row=header_row + i, column=5).protection = \
            __import__("openpyxl.styles.protection", fromlist=["Protection"]).Protection(locked=False)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"External_Template_{_sanitize(meta['subject_code'])}.xlsx"
    return buf, filename

def parse_external_upload(file_obj, subject_code, semester_id, academic_year):
    try:
        wb = load_workbook(file_obj, data_only=True)
        ws = wb.active
    except Exception as e:
        return False, [{'row': 0, 'col': 'FILE', 'msg': f'Cannot read file: {str(e)}'}], 0

    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and 'Roll No' in str(cell.value):
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        return False, [{'row': 0, 'col': 'STRUCTURE', 'msg': 'Header row not found. Use the official template.'}], 0

    # Validate template structure
    header_vals = [ws.cell(row=header_row, column=c).value for c in range(1, 6)]
    expected_headers = ["SR", "Roll No", "PRN", "Student Name"]
    for i, eh in enumerate(expected_headers):
        if not header_vals[i] or eh.lower() not in str(header_vals[i]).lower():
            return False, [{'row': header_row, 'col': 'STRUCTURE', 'msg': f'Invalid column header. Expected column {i+1} to be "{eh}".'}], 0

    valid_prns = {s.prn for s in Student.query.all()}
    
    # Pre-fetch all enrolled PRNs for this subject
    enrolled_prns = {
        e.prn for e in Enrollment.query.filter_by(
            subject_code=subject_code,
            semester_id=semester_id,
            academic_year=academic_year
        ).all()
    }

    ABSENT_VALUES = {'ab', 'absent', 'a', ''}
    errors = []
    records = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row[2]:  # PRN is index 2
            continue
        
        prn_val = str(row[2]).strip()
        marks_val = row[4]  # Marks is index 4

        if prn_val.lower().startswith('note'):
            continue

        row_num = header_row + 1 + len(records) + len(errors)

        # 1. PRN exists validation
        if prn_val not in valid_prns:
            errors.append({'row': row_num, 'col': 'PRN', 'msg': f'PRN {prn_val} not found in database'})
            continue

        # 2. Student is enrolled in selected subject validation
        if prn_val not in enrolled_prns:
            errors.append({'row': row_num, 'col': 'Enrollment', 'msg': f'Student with PRN {prn_val} is not enrolled in this subject'})
            continue

        if marks_val is None:
            records.append({'prn': prn_val, 'external_marks': None, 'absent': True})
            continue

        str_val = str(marks_val).strip().lower()
        if str_val in ABSENT_VALUES:
            records.append({'prn': prn_val, 'external_marks': None, 'absent': True})
            continue

        # 3. Marks are numeric validation
        try:
            marks = float(str_val)
        except (ValueError, TypeError):
            errors.append({'row': row_num, 'col': 'External Marks', 'msg': f'Value "{marks_val}" is not a valid number.'})
            continue

        # 4. Marks are between 0 and 60 validation
        if not (0 <= marks <= 60):
            errors.append({'row': row_num, 'col': 'External Marks', 'msg': f'External marks {marks} out of range. Must be 0 to 60.'})
            continue

        records.append({'prn': prn_val, 'external_marks': marks, 'absent': False})

    if errors:
        return False, errors, 0

    return True, records, len(records)

def generate_external_excel_report(meta, data):
    """External marks Excel report with the same logo/college header style."""
    wb = Workbook()
    ws = wb.active
    ws.title = "External Report"

    last_col = 9
    last_letter = "I"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14

    # ===== COMMON LOGO HEADER =====
    if os.path.exists(LEFT_LOGO):
        left_logo = Image(LEFT_LOGO)
        left_logo.width = 85
        left_logo.height = 85
        ws.add_image(left_logo, "A1")

    if os.path.exists(RIGHT_LOGO):
        right_logo = Image(RIGHT_LOGO)
        right_logo.width = 85
        right_logo.height = 85
        ws.add_image(right_logo, "I1")

    ws.merge_cells("B1:H1")
    c = ws["B1"]
    c.value = (
        "Chhatrapati Shahu Maharaj Shikshan Sanstha's\n"
        "CSMSS Chh. Shahu College of Engineering\n"
        "Kanchanwadi, Chhatrapati Sambhajinagar – 431011\n"
        "Department of Electronics and Computer Engineering"
    )
    c.font = Font(name=TNR, size=13, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _thick_border()
    ws.row_dimensions[1].height = 95
    ws.row_dimensions[2].height = 6

    for row in range(1, 2):
        for col in range(1, last_col + 1):
            ws.cell(row=row, column=col).border = _thick_border()

    # ===== REPORT TITLE =====
    ws.merge_cells("A3:I3")
    ws["A3"] = "External Marks Report" + (" — " + meta.get("batch_label", "All Students") if meta.get("batch_label") else "")
    ws["A3"].font = Font(name=TNR, size=12, bold=True)
    ws["A3"].alignment = _center()
    ws["A3"].border = _thick_border()

    ws.merge_cells("A4:B4")
    ws["A4"] = "Subject Code :"
    ws.merge_cells("C4:D4")
    ws["C4"] = meta['subject_code']
    ws.merge_cells("E4:F4")
    ws["E4"] = "Subject Name :"
    ws.merge_cells("G4:I4")
    ws["G4"] = meta['subject_name']

    ws.merge_cells("A5:B5")
    ws["A5"] = "Semester :"
    ws.merge_cells("C5:D5")
    ws["C5"] = meta['semester_id']
    ws.merge_cells("E5:F5")
    ws["E5"] = "Academic Year :"
    ws.merge_cells("G5:I5")
    ws["G5"] = meta['academic_year']

    for r in range(4, 6):
        for cidx in range(1, last_col + 1):
            cell = ws.cell(r, cidx)
            cell.border = _thick_border()
            cell.font = _body_font(10, bold=cidx in [1, 5])
            cell.alignment = _left() if cidx in [1, 5] else _center()

    # ===== TABLE =====
    hdrs = [
        "SR", "PRN", "Student Name", "Internal (/40)", "External (/60)",
        "Total (/100)", "Grade", "Grade Point", "Result"
    ]
    hr = 7
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    pass_fill = PatternFill("solid", fgColor="D5F5E3")
    fail_fill = PatternFill("solid", fgColor="FADBD8")
    pending_fill = PatternFill("solid", fgColor="F2F3F4")

    for cidx, h in enumerate(hdrs, 1):
        cell = ws.cell(row=hr, column=cidx, value=h)
        cell.font = Font(name=TNR, bold=True, size=10)
        cell.fill = header_fill
        cell.border = _thick_border()
        cell.alignment = _center()

    pass_c = fail_c = pending_c = 0
    marks_list = []

    for i, s in enumerate(data, 1):
        row = hr + i
        ext = s["external"]
        internal = s["internal"]
        total = s["total"]
        status = s["status"]

        fill = pass_fill if status == "PASS" else (fail_fill if status == "FAIL" else pending_fill)

        values = [
            i,
            s["prn"],
            s["name"],
            internal if internal is not None else "—",
            ext if ext is not None else "—",
            total if total is not None else "—",
            s["grade"] or "—",
            s["grade_point"] if s["grade_point"] is not None else "—",
            status,
        ]

        for cidx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=cidx, value=value)
            cell.font = _body_font(10)
            cell.alignment = _center() if cidx != 3 else _left()
            cell.border = _thin_border()
            cell.fill = fill

        if ext is not None:
            marks_list.append(ext)
        if status == "PASS":
            pass_c += 1
        elif status == "FAIL":
            fail_c += 1
        else:
            pending_c += 1

    # ===== SUMMARY =====
    sr = hr + len(data) + 2
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=9)
    ws.cell(row=sr, column=1).value = (
        f"Summary: Total {len(data)} | Entered {len(marks_list)} | "
        f"Pass {pass_c} | Fail {fail_c} | Pending {pending_c} | "
        f"Pass% {round(pass_c/len(marks_list)*100,1) if marks_list else 0}% | "
        f"Highest {max(marks_list) if marks_list else '—'} | "
        f"Lowest {min(marks_list) if marks_list else '—'} | "
        f"Average {round(sum(marks_list)/len(marks_list),2) if marks_list else '—'}"
    )
    ws.cell(row=sr, column=1).font = Font(name=TNR, bold=True, size=10)
    ws.cell(row=sr, column=1).alignment = _center()
    ws.cell(row=sr, column=1).border = _thick_border()

    # ===== SIGNATURES =====
    sig_row = sr + 3
    ws.row_dimensions[sig_row].height = 60
    ws.merge_cells(f"A{sig_row}:C{sig_row}")
    ws.merge_cells(f"D{sig_row}:F{sig_row}")
    ws.merge_cells(f"G{sig_row}:I{sig_row}")

    ws[f"A{sig_row}"] = "Subject Teacher"
    ws[f"D{sig_row}"] = "HOD"
    ws[f"G{sig_row}"] = "Principal"

    for cell_ref in [f"A{sig_row}", f"D{sig_row}", f"G{sig_row}"]:
        cell = ws[cell_ref]
        cell.font = Font(name=TNR, size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
        cell.border = _thick_border()

    footer_row = sig_row + 2
    ws.merge_cells(f"A{footer_row}:I{footer_row}")
    ws[f"A{footer_row}"] = "Generated by ECE Student Management System — CSMSS Chh. Shahu College of Engineering"
    ws[f"A{footer_row}"].font = Font(name=TNR, size=9, italic=True)
    ws[f"A{footer_row}"].alignment = _center()

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"External_Report_{meta['subject_code']}_Sem{meta['semester_id']}.xlsx"
    return output, filename
